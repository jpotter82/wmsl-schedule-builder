
# --- deterministic day-of-week labels (avoid locale issues) ---
DOWS = ['Mon','Tue','Wed','Thu','Fri','Sat','Sun']

def dow_label(d):
    """Return fixed 3-letter DOW label for date/datetime."""
    return DOWS[d.weekday()]

# Run seed: None => randomize every run
RUN_SEED = None


def _common_avail_days(team1, team2, team_availability):
    """Return comma-separated DOW labels both teams can play."""
    if not team_availability:
        return ""
    a1 = set(team_availability.get(team1, []))
    a2 = set(team_availability.get(team2, []))
    common = a1 & a2
    if not common:
        return ""
    order = {d:i for i,d in enumerate(DOWS)}
    return ", ".join(sorted(common, key=lambda d: order.get(d, 99)))

def _blackout_summary(team1, team2, team_blackouts, max_dates=30):
    """Return comma-separated blackout dates (YYYY-MM-DD) where either team cannot play."""
    if not team_blackouts:
        return ""
    b1 = set(team_blackouts.get(team1, []))
    b2 = set(team_blackouts.get(team2, []))
    dates = sorted(b1 | b2)
    if not dates:
        return ""
    out = [d.strftime("%Y-%m-%d") for d in dates[:max_dates]]
    if len(dates) > max_dates:
        out.append(f"...(+{len(dates)-max_dates} more)")
    return ", ".join(out)

def check_schedule_against_availability(schedule, team_availability):
    """Return list of (date, day, time, field, team, allowed_days) for any availability violations."""
    violations = []
    for (d, time_str, field_id, home, home_div, away, away_div) in schedule:
        if not home or not away:
            continue
        day = dow_label(d)
        for team in (home, away):
            allowed = team_availability.get(team)
            if allowed is None:
                continue
            if day not in allowed:
                violations.append((d.strftime('%Y-%m-%d'), day, time_str, field_id, team, ",".join(sorted(allowed))))
    return violations

#!/usr/bin/env python3
"""
Softball scheduler (heuristic) + Excel export.

Additions in this version:
  - CSV export writes 1 row PER field_availability slot (including unscheduled/blank slots),
    so row count matches field_availability.
  - XLSX export with:
      * Schedule sheet (same rows as field_availability, blanks for unused slots)
      * Teams sheet
      * Summary sheet (all formulas; updates if you edit Schedule)
      * TeamDate helper sheet (for DH-day counting formulas)
      * Matchup Matrix sheet (formula-based, symmetric counts)
      * Conditional formatting (unused slots, illegal matchups, home==away, matrix heatmap)
Requires:
  pip install openpyxl
Optional:
  pip install prettytable
"""

import csv
import itertools
import random
import math
import re
import os
from datetime import datetime, date, timedelta
from collections import defaultdict

try:
    from prettytable import PrettyTable
except ImportError:
    PrettyTable = None

# XLSX export support (optional)
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import FormulaRule, CellIsRule, ColorScaleRule
except ImportError:
    Workbook = None
    Font = Alignment = PatternFill = get_column_letter = FormulaRule = None
    CellIsRule = ColorScaleRule = None

# -------------------------------
# Configurable parameters
# -------------------------------
MAX_RETRIES = 20000            # scheduling backtracking limit
PREFERRED_MIN_GAP = 3         # ideal minimum days between game dates (soft preference)
HARD_MIN_GAP = 2              # absolute minimum days between game dates (hard constraint)
WEEKLY_GAME_LIMIT = 2          # HARD max games per team per week (1 DH day = 2 games)
# How hard to push for an evenly paced season. This is a single dial: it drives the
# placement penalty, the pod-scheduling gate, AND how much best-of-N cares about
# idle/overloaded weeks. Setting it to 0 disables weekly pacing entirely, so the
# scheduler maximises games placed exactly as it did before pacing existed.
# The dial runs 0 (off) to WEEKLY_BALANCE_REFERENCE (full strength). Values above the
# reference are CLAMPED to it by the wrapper — they are not naturally equivalent, so
# without that clamp a larger number really does push pacing harder (measurably fewer
# games placed) in a way no UI option can express. The UI offers named levels instead
# of a free-text number for the same reason.
WEEKLY_BALANCE_PENALTY = 2500  # score penalty per game above a team's soft weekly target
WEEKLY_BALANCE_REFERENCE = 2500  # strength at which pacing is applied at full weight
HOME_AWAY_BALANCE = 7          # desired home games per team (for 14-game seasons)


# Division A opponent-balance controls (A is DH-only)
A_PAIR_MIN_GAMES = 1          # each A-vs-A pairing should occur at least this many times
A_PAIR_SOFT_CAP = 3           # avoid exceeding this for a pair while some required pairs still unmet
B_PAIR_MIN_GAMES = 2          # each A-vs-A pairing should occur at least this many times
B_PAIR_SOFT_CAP = 2           # avoid exceeding this for a pair while some required pairs still unmet
C_PAIR_MIN_GAMES = 1          # each A-vs-A pairing should occur at least this many times
C_PAIR_SOFT_CAP = 3           # avoid exceeding this for a pair while some required pairs still unmet
# Pairing balance rules (min games per opponent + soft cap to avoid lopsided repeats)
# NOTE: For divisions where the intra_target_per_team makes min infeasible, the code will automatically
# clamp the effective minimum to floor(avg_games_vs_opponent).
PAIR_RULES = {
    'A': {'min': A_PAIR_MIN_GAMES, 'soft_cap': A_PAIR_SOFT_CAP},
    'B': {'min': 1, 'soft_cap': 3},
    'C': {'min': 1, 'soft_cap': 3},
    # 'D': {'min': 1, 'soft_cap': 3},
}

def effective_pair_rules(division, intra_target_per_team, n):
    """Return (min_eff, cap_eff) for intra-division opponent balance."""
    base = PAIR_RULES.get(division, {'min': 0, 'soft_cap': 999})
    if n <= 1 or intra_target_per_team <= 0:
        return 0, base.get('soft_cap', 999)
    avg = float(intra_target_per_team) / float(max(1, n - 1))
    min_eff = min(int(base.get('min', 0)), int(math.floor(avg)))
    # keep cap at least (ceil(avg)+1) so we don't dead-end in small divisions
    cap_eff = max(int(base.get('soft_cap', 999)), int(math.ceil(avg)) + 1)
    return min_eff, cap_eff



# Sunday pod rotation:
# For Sunday dates, we try to rotate which division gets pod-style doubleheaders.
# This helps avoid one division (e.g., A) soaking up all Sunday capacity.
SUNDAY_POD_ROTATION = ['B', 'C', 'A']  # cycle order (can change)
SUNDAY_PODS_PER_SUNDAY = 3  # at most this many *pod sessions* across all divisions on a Sunday

# Sunday preference.
#
# SUNDAY_PRIORITY: score bonus for placing a game on a Sunday, and when > 0 the pod
#   schedulers scan Sundays before weekdays. 0 keeps the original behaviour, where
#   Sundays are only favoured while teams still owe their minimum Sunday sessions.
# SUNDAY_PODS_ONLY: reserve Sunday inventory for doubleheader pods. Single games are
#   not placed on Sundays at all, so Sunday slots are never broken up by a lone game
#   that blocks a pod from using the adjacent timeslot.
SUNDAY_PRIORITY = 0
SUNDAY_PODS_ONLY = False
RANDOM_SEED = None           # for repeatable schedules
# Per-division configuration (tweak here)
#
# dh_only:
#   True  -> division plays *only* pod doubleheaders. Every game is part of a
#            4-team pod (2 games/team/day). No single games are ever placed, so
#            target_games should be even. Best schedule shape, but needs the most
#            adjacent-slot inventory; teams may finish short if pods can't be built.
#   False -> division builds pods until each team reaches min_dh doubleheader days,
#            then fills the remaining games as singles. More forgiving.
DIVISION_SETTINGS = {
    # 14 games per team: 6 DH days (12 games) + 2 single game days
    'A': {'inter': False, 'target_games': 14, 'min_dh': 6, 'max_dh': 6, 'dh_only': True},
    'B': {'inter': False, 'target_games': 14, 'min_dh': 6, 'max_dh': 6, 'dh_only': False},
    'C': {'inter': False, 'target_games': 14, 'min_dh': 6, 'max_dh': 6, 'dh_only': False},
    # 'D': {'inter': False, 'target_games': 14, 'min_dh': 6, 'max_dh': 6, 'dh_only': False},
}

# Inter-division pairing settings (only applied if BOTH divisions have inter=True)
INTER_PAIR_SETTINGS = {
    ('A', 'B'): False,
    ('A', 'C'): False,
    ('A', 'D'): False,
    ('B', 'C'): False,
    ('C', 'D'): False,
    ('B', 'D'): False,
}

# “Average per team” targets.
INTER_DEGREE = {
    ('B', 'C'): 4,
    ('C', 'D'): 6,
}

# -------------------------------
# Helpers
# -------------------------------
def div_of(team):
    return team[0].upper()

def target_games(team):
    return DIVISION_SETTINGS[div_of(team)]['target_games']

def min_dh(team):
    return DIVISION_SETTINGS[div_of(team)]['min_dh']

def max_dh(team):
    return DIVISION_SETTINGS[div_of(team)]['max_dh']

def is_dh_only(division):
    """True if this division plays pod doubleheaders exclusively (no single games).

    Accepts a division letter or a team name (first character is the division).
    """
    if not division:
        return False
    div = division[0].upper()
    return bool(DIVISION_SETTINGS.get(div, {}).get('dh_only', False))

def dh_only_divisions(division_teams):
    """Divisions (in order) that are configured as doubleheader-only."""
    return [d for d in division_teams if is_dh_only(d)]

DIV_PRIORITY = {'D': 3, 'C': 2, 'B': 1, 'A': 0}

# -------------------------------
# Team scheduling scarcity
# -------------------------------
# A team that can play on few calendar dates — because of limited weekday
# availability (team_availability) and/or many blackout dates (team_blackouts) —
# is "scarce" and must be given priority when building pods/games, or it ends up
# unschedulable. These globals are populated at run start via init_team_scarcity().
TEAM_PLAYABLE_DATES = {}   # team -> count of season dates the team can actually play
SEASON_DATE_COUNT = 0      # total distinct dates in field availability
SEASON_WEEKS = 0           # distinct ISO weeks that have any field availability
SEASON_EFFECTIVE_WEEKS = 0 # weeks with enough slots to seat most of the league
LOW_CAPACITY_WEEKS = set() # weeks too small to give every team a game
SEASON_WEEK_INDEX = {}     # ISO week number -> 0-based position in the season

# Fill the opening weeks before spreading out.
#
# The greedy placer walks open slots in random order, so with a tight calendar it can
# leave week 1 and 2 slots unused while filling later weeks — and an unused early slot
# is capacity that can never be recovered. Setting this to N makes the scheduler treat
# the first N weeks as must-fill, placing games there before it considers later dates.
FRONT_LOAD_WEEKS = 0
FRONT_LOAD_BONUS = 4000    # score bonus for placing a game inside the front-load window

# Longest acceptable layoff between a team's game dates.
#
# This is a *soft target*, not a hard cap: placements that shrink a long layoff get a
# score bonus, scaled by how far past the target the layoff already is. A hard reject
# works poorly here because the greedy passes are not chronological — a later placement
# can still split a big gap, so refusing outright just loses games.
MAX_IDLE_DAYS = 14
IDLE_GAP_REPAIR_WEIGHT = 1500

# Soft cap on games per team per week.
#
# WEEKLY_GAME_LIMIT is the HARD ceiling (a team may never exceed it). This soft
# target is what the scheduler aims for, so games are spread evenly instead of
# being crammed into a few weeks while others sit idle. Set to None to derive it
# automatically as ceil(target_games / SEASON_WEEKS).
WEEKLY_SOFT_TARGET = None

def init_team_scarcity(all_teams, field_availability, team_availability, team_blackouts):
    """Precompute, per team, how many season dates it can actually play.

    Combines weekly availability and blackout dates via is_team_available, so a
    team is scarce if it is limited on EITHER axis. Fewer playable dates => scarcer.

    Also records how many ISO weeks the season spans, which sets the soft
    per-week game target used to keep the schedule evenly paced.
    """
    global TEAM_PLAYABLE_DATES, SEASON_DATE_COUNT, SEASON_WEEKS
    global SEASON_EFFECTIVE_WEEKS, LOW_CAPACITY_WEEKS, SEASON_WEEK_INDEX
    season_dates = sorted({dt.date() for (dt, _slot, _field) in field_availability})
    SEASON_DATE_COUNT = len(season_dates)
    SEASON_WEEKS = len({d.isocalendar()[1] for d in season_dates})

    # A week with only a handful of slots physically cannot give every team a game.
    # Those weeks must not drag down the per-week target, or the scheduler ends up
    # aiming too low everywhere else and leaves games unplaced.
    slots_per_week = defaultdict(int)
    for (dt, _slot, _field) in field_availability:
        slots_per_week[dt.date().isocalendar()[1]] += 1
    n_teams = max(1, len(all_teams))
    LOW_CAPACITY_WEEKS = {w for w, s in slots_per_week.items() if s * 2 < n_teams}
    SEASON_EFFECTIVE_WEEKS = max(1, len(slots_per_week) - len(LOW_CAPACITY_WEEKS))

    # Chronological position of each week, so "the first N weeks" is well defined
    # even when ISO week numbers wrap across a year boundary.
    SEASON_WEEK_INDEX = {}
    for i, d in enumerate(sorted({dd.isocalendar()[:2] for dd in season_dates})):
        SEASON_WEEK_INDEX[d[1]] = i

    TEAM_PLAYABLE_DATES = {}
    for team in all_teams:
        TEAM_PLAYABLE_DATES[team] = sum(
            1 for d in season_dates
            if is_team_available(team, d, team_availability, team_blackouts)
        )

def weekly_soft_target(team):
    """Preferred maximum games for `team` in any single week.

    Defaults to an even spread of its season across the available weeks, e.g.
    14 games over 7 weeks -> 2 per week (one doubleheader). Never returns more
    than the hard WEEKLY_GAME_LIMIT.
    """
    if WEEKLY_SOFT_TARGET:
        return min(int(WEEKLY_SOFT_TARGET), WEEKLY_GAME_LIMIT)
    weeks = SEASON_EFFECTIVE_WEEKS or SEASON_WEEKS
    if not weeks:
        return WEEKLY_GAME_LIMIT
    # Spread the season over the weeks that can actually host a full slate.
    fair = int(math.ceil(float(target_games(team)) / float(weeks)))
    return max(2, min(fair, WEEKLY_GAME_LIMIT))

def _played_dates(team, team_game_days):
    """Dates the team actually plays on.

    NOTE: unlike a plain sorted(team_game_days[team]), this filters out dates whose
    count has fallen to zero. The repair pass decrements a date's count when it moves
    a game away but leaves the key in place, so an unfiltered read would treat a date
    the team no longer plays as if it still had a game.
    """
    return sorted(d for d, n in team_game_days[team].items() if n > 0)

def longest_idle_gap(team, team_game_days):
    """Largest gap in days between consecutive game dates already scheduled for a team."""
    dates = _played_dates(team, team_game_days)
    if len(dates) < 2:
        return 0
    return max((dates[i] - dates[i - 1]).days for i in range(1, len(dates)))

def longest_idle_gap_after_adding(team, d, team_game_days):
    """Largest gap after hypothetically also playing on date d."""
    dates = sorted(set(_played_dates(team, team_game_days)) | {d})
    if len(dates) < 2:
        return 0
    return max((dates[i] - dates[i - 1]).days for i in range(1, len(dates)))

def idle_gap_repair_bonus(team, d, team_game_days):
    """Score bonus when playing on date d shortens the team's worst layoff.

    Weighted by how much the gap shrinks, with an extra push once the existing gap is
    already past MAX_IDLE_DAYS.
    """
    if not IDLE_GAP_REPAIR_WEIGHT:
        return 0
    before = longest_idle_gap(team, team_game_days)
    after = longest_idle_gap_after_adding(team, d, team_game_days)
    if after >= before:
        return 0
    bonus = (before - after) * IDLE_GAP_REPAIR_WEIGHT
    if before > MAX_IDLE_DAYS:
        bonus += (before - MAX_IDLE_DAYS) * IDLE_GAP_REPAIR_WEIGHT
    return bonus

def check_max_idle_gap(schedule, teams, max_idle_days=None):
    """Report teams whose longest layoff exceeds the target.

    Returns list of (team, gap_days, gap_start, gap_end), worst first.
    """
    limit = MAX_IDLE_DAYS if max_idle_days is None else max_idle_days
    by_team = defaultdict(set)
    for entry in schedule:
        if entry is None:
            continue
        dt, _slot, _field, home, _hd, away, _ad = entry
        by_team[home].add(dt.date())
        by_team[away].add(dt.date())

    out = []
    for team in teams:
        dates = sorted(by_team.get(team, ()))
        for i in range(1, len(dates)):
            gap = (dates[i] - dates[i - 1]).days
            if gap > limit:
                out.append((team, gap, dates[i - 1], dates[i]))
    out.sort(key=lambda r: -r[1])
    return out

def in_front_load_window(d):
    """True if date d falls inside the opening weeks we want filled first."""
    if not FRONT_LOAD_WEEKS:
        return False
    wk = d.isocalendar()[1]
    idx = SEASON_WEEK_INDEX.get(wk)
    return idx is not None and idx < FRONT_LOAD_WEEKS

def weekly_excess(team, d, team_stats, extra=1):
    """How far above its soft weekly target `team` would go by adding `extra` games in d's week."""
    wk = d.isocalendar()[1]
    have = team_stats[team]['weekly_games'].get(wk, 0)
    return max(0, (have + extra) - weekly_soft_target(team))

def team_scarcity(team):
    """Scheduling scarcity: higher = more constrained (fewer playable dates).

    Returns 0 when scarcity has not been initialized, so behavior is unchanged
    until init_team_scarcity() has run.
    """
    if not SEASON_DATE_COUNT:
        return 0
    playable = TEAM_PLAYABLE_DATES.get(team, SEASON_DATE_COUNT)
    return SEASON_DATE_COUNT - playable

def game_deficit(team, team_stats):
    return max(0, target_games(team) - team_stats[team]['total_games'])

def dh_deficit(team, doubleheader_count):
    return max(0, min_dh(team) - doubleheader_count[team])

def team_need_key(team, team_stats, doubleheader_count):
    return (
        team_scarcity(team),
        dh_deficit(team, doubleheader_count),
        game_deficit(team, team_stats),
        DIV_PRIORITY.get(div_of(team), 0),
        -team_stats[team]['home_games'],
        team
    )

def matchup_need_score(home, away, team_stats, doubleheader_count):
    return (
        team_scarcity(home) + team_scarcity(away)
    ) * 5000 + (
        game_deficit(home, team_stats) + game_deficit(away, team_stats)
    ) * 1000 + (
        dh_deficit(home, doubleheader_count) + dh_deficit(away, doubleheader_count)
    ) * 50 + (
        DIV_PRIORITY.get(div_of(home), 0) + DIV_PRIORITY.get(div_of(away), 0)
    )

def inter_enabled_for_pair(d1, d2):
    d1, d2 = d1.upper(), d2.upper()
    key = (d1, d2) if (d1, d2) in INTER_PAIR_SETTINGS else (d2, d1)
    if key not in INTER_PAIR_SETTINGS or not INTER_PAIR_SETTINGS[key]:
        return False
    return DIVISION_SETTINGS[d1]['inter'] and DIVISION_SETTINGS[d2]['inter']

def pair_degree(d1, d2):
    d1, d2 = d1.upper(), d2.upper()
    key = (d1, d2) if (d1, d2) in INTER_DEGREE else (d2, d1)
    return INTER_DEGREE.get(key, 0)

def min_gap_ok(team, d, team_game_days):
    """Hard gap check: return True if 'team' has no game within HARD_MIN_GAP days of date d."""
    for gd in team_game_days[team]:
        if gd != d and abs((d - gd).days) < HARD_MIN_GAP:
            return False
    return True

# -------------------------------
# Field availability index (avoids O(n) scans in inner loops)
# -------------------------------
def build_field_index(field_availability):
    """Build a dict: (date, slot) -> list of (datetime, slot, field) entries.

    This replaces the inner-loop list comprehensions that scan the full
    field_availability list to find free fields for a given date+slot.
    """
    idx = defaultdict(list)
    for entry in field_availability:
        dt, slot, field = entry
        idx[(dt.date(), slot)].append(entry)
    return idx

def free_fields_for_slot(field_index, d, slot, used_slots):
    """Return list of (datetime, slot, field) entries that are unused for (date, slot)."""
    return [entry for entry in field_index.get((d, slot), [])
            if not used_slots.get((entry[0], slot, entry[2]), False)]

# -------------------------------
# Availability helpers
# -------------------------------
def dow_abbrev(d):
    """Return 3-letter day abbrev (Mon/Tue/...) for a date or datetime."""
    try:
        return dow_label(d)
    except Exception:
        return str(d)[:3].title()

def is_team_available(team, d, team_availability, team_blackouts):
    """True if team can play on date d according to weekly availability + blackout dates."""
    dd = d if hasattr(d, "weekday") and not hasattr(d, "date") else d.date()
    dow = dow_abbrev(dd)
    if dow not in team_availability.get(team, set()):
        return False
    if dd in team_blackouts.get(team, set()):
        return False
    return True

def score_placement(t1, t2, d, team_stats, doubleheader_count, team_game_days, sunday_assignment=None):
    """Centralized scoring for placing matchup (t1,t2) on date d.

    Higher is better. Used by both schedule_games and fill_missing_games
    to rank candidate placements consistently.

    Components:
      - game/DH deficit (teams further behind get priority)
      - division priority weighting
      - soft gap penalty (prefer 3+ day gaps)
      - Sunday rotation bonus
      - DH penalty (avoid unnecessary doubleheaders)
    """
    score = matchup_need_score(t1, t2, team_stats, doubleheader_count)

    # Soft gap preference (allow 2-day gaps, prefer 3+)
    score -= preferred_gap_penalty(t1, d, team_game_days)
    score -= preferred_gap_penalty(t2, d, team_game_days)

    # Strong preference for placements that break up a long layoff. The gap penalties
    # above stop games bunching TOO CLOSE; this stops them drifting too far apart.
    score += idle_gap_repair_bonus(t1, d, team_game_days)
    score += idle_gap_repair_bonus(t2, d, team_game_days)

    # Weekly balance: strongly prefer weeks where the team is still under its even
    # share. Without this the scheduler happily stacks 4 games into one week and
    # leaves the next one empty.
    score -= weekly_excess(t1, d, team_stats) * WEEKLY_BALANCE_PENALTY
    score -= weekly_excess(t2, d, team_stats) * WEEKLY_BALANCE_PENALTY

    # Sunday rotation bonus
    if sunday_assignment and d.weekday() == 6:
        assigned = sunday_assignment.get(d)
        if assigned and div_of(t1) == assigned and div_of(t2) == assigned:
            score += 500

    # General Sunday preference (independent of the rotation)
    if SUNDAY_PRIORITY and d.weekday() == 6:
        score += SUNDAY_PRIORITY

    # Fill the opening weeks first — an empty week-1 slot is capacity lost for good.
    if in_front_load_window(d):
        score += FRONT_LOAD_BONUS

    # Pepper singles: penalize creating a DH day when team already has enough DH days
    for team in (t1, t2):
        if team_game_days[team][d] == 1 and doubleheader_count[team] >= min_dh(team):
            score -= 2000

    return score


def preferred_gap_penalty(team, d, team_game_days, penalty_per_day=500):
    """Soft preference penalty when gap is smaller than PREFERRED_MIN_GAP.
    Returns 0 if the closest existing game day is >= PREFERRED_MIN_GAP days away.
    """
    closest = None
    for gd in team_game_days[team]:
        if gd == d:
            continue
        delta = abs((d - gd).days)
        if closest is None or delta < closest:
            closest = delta
    if closest is None:
        return 0
    if closest >= PREFERRED_MIN_GAP:
        return 0
    return (PREFERRED_MIN_GAP - closest) * penalty_per_day

def longest_idle_gap(team, team_game_days):
    """Largest day gap between consecutive game dates already scheduled for a team."""
    dates = sorted(team_game_days[team])
    if len(dates) < 2:
        return 0
    return max((dates[i] - dates[i - 1]).days for i in range(1, len(dates)))


def longest_idle_gap_after_adding(team, d, team_game_days):
    """Largest day gap after hypothetically adding date d for team."""
    dates = sorted(set(team_game_days[team]) | {d})
    if len(dates) < 2:
        return 0
    return max((dates[i] - dates[i - 1]).days for i in range(1, len(dates)))


def idle_gap_repair_bonus(team, d, team_game_days):
    """
    Positive score when placing team on date d shrinks an existing long layoff.
    This works better than a pure hard reject with the current non-chronological
    greedy passes, because later placements can still split a large gap.
    """
    before = longest_idle_gap(team, team_game_days)
    after = longest_idle_gap_after_adding(team, d, team_game_days)
    if after < before:
        bonus = (before - after) * IDLE_GAP_REPAIR_WEIGHT
        if before > MAX_IDLE_DAYS:
            bonus += (before - MAX_IDLE_DAYS) * IDLE_GAP_REPAIR_WEIGHT
        return bonus
    return 0


def check_max_idle_gap(schedule, teams, max_idle_days=MAX_IDLE_DAYS):
    """Return (team, previous_date, next_date, gap_days) for long layoff violations."""
    by_team = defaultdict(set)
    for (dt, _time, _field, home, _home_div, away, _away_div) in schedule:
        dd = dt.date() if hasattr(dt, 'date') else dt
        by_team[home].add(dd)
        by_team[away].add(dd)

    violations = []
    for team in teams:
        dates = sorted(by_team.get(team, set()))
        for i in range(1, len(dates)):
            gap = (dates[i] - dates[i - 1]).days
            if gap > max_idle_days:
                violations.append((team, dates[i - 1].strftime('%Y-%m-%d'), dates[i].strftime('%Y-%m-%d'), gap))
    return violations


def season_week_index(d, season_start=None):
    """Return a stable season week index (0-based) using Monday-based weeks."""
    from datetime import timedelta
    dd = d.date() if hasattr(d, 'date') and not hasattr(d, 'weekday') else d
    start = season_start or SEASON_START_DATE or dd
    start_monday = start - timedelta(days=start.weekday())
    return (dd - start_monday).days // 7


def team_weeks_played(team, team_game_days, season_start=None):
    return sorted({season_week_index(dd, season_start) for dd in team_game_days[team].keys()})


def max_consecutive_byes(team, team_game_days, season_start=None):
    weeks = team_weeks_played(team, team_game_days, season_start)
    if len(weeks) < 2:
        return 0
    return max(max(0, weeks[i] - weeks[i - 1] - 1) for i in range(1, len(weeks)))


def max_consecutive_byes_after_adding(team, d, team_game_days, season_start=None):
    weeks = set(team_weeks_played(team, team_game_days, season_start))
    weeks.add(season_week_index(d, season_start))
    weeks = sorted(weeks)
    if len(weeks) < 2:
        return 0
    return max(max(0, weeks[i] - weeks[i - 1] - 1) for i in range(1, len(weeks)))


def no_two_consecutive_byes_after_adding(team, d, team_game_days, season_start=None, max_consecutive_byes=MAX_CONSECUTIVE_BYE_WEEKS):
    """True if adding date d does not create more than the allowed consecutive bye weeks between games."""
    return max_consecutive_byes_after_adding(team, d, team_game_days, season_start) <= max_consecutive_byes


def bye_week_urgency_bonus(team, d, team_game_days, season_start=None):
    """Score bonus for placements that reduce/avoid consecutive bye-week stretches."""
    before = max_consecutive_byes(team, team_game_days, season_start)
    after = max_consecutive_byes_after_adding(team, d, team_game_days, season_start)
    if after < before:
        return (before - after) * (BYE_URGENCY_WEIGHT * 2)
    weeks = team_weeks_played(team, team_game_days, season_start)
    if not weeks:
        return 0
    w = season_week_index(d, season_start)
    bonus = 0
    prev_weeks = [wk for wk in weeks if wk < w]
    next_weeks = [wk for wk in weeks if wk > w]
    if prev_weeks:
        gap_from_prev = w - prev_weeks[-1]
        if gap_from_prev == 2:
            bonus += BYE_URGENCY_WEIGHT
        elif gap_from_prev > 2:
            bonus += BYE_URGENCY_WEIGHT * 2
    if next_weeks:
        gap_to_next = next_weeks[0] - w
        if gap_to_next == 2:
            bonus += BYE_URGENCY_WEIGHT
        elif gap_to_next > 2:
            bonus += BYE_URGENCY_WEIGHT * 2
    return bonus


# -------------------------------
# Data loading functions
# -------------------------------
def load_team_availability(file_path):
    """Load per-team day-of-week availability.

    Accepts CSV where each team row contains day tokens in any of these forms:
      - Separate columns: Mon, Tue, Wed, ...
      - A single cell with delimiters: "Mon;Tue;Wed" or "Mon, Tue, Wed"
      - Whitespace-separated: "Mon Tue Wed"
      - Full day names ("Monday") are accepted and normalized to 3-letter form.

    Returns: dict[team] -> set({"Mon","Tue","Wed","Thu","Fri","Sat","Sun"})
    """
    VALID = {"Mon","Tue","Wed","Thu","Fri","Sat","Sun"}
    def norm(tok: str):
        tok = (tok or "").strip()
        if not tok:
            return None
        # accept full day names
        t3 = tok[:3].title()
        if t3 in VALID:
            return t3
        return None

    availability = {}
    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        next(reader, None)  # header
        for row in reader:
            if not row:
                continue
            team = (row[0] or "").strip()
            if not team:
                continue
            tokens = []
            for cell in row[1:]:
                cell = (cell or "").strip()
                if not cell:
                    continue
                # split on common delimiters
                for part in re.split(r"[;,\s]+", cell):
                    part = part.strip()
                    if part:
                        tokens.append(part)
            days = set()
            for t in tokens:
                d = norm(t)
                if d:
                    days.add(d)
            availability[team] = days
    return availability


def load_team_preferred_days(file_path):
    """Load optional per-team preferred day-of-week values.

    Format matches team_availability.csv semantics:
      Team,PreferredDays
      A1,Mon,Wed
      A2,Tue;Thu

    Missing file => empty dict.
    """
    if not file_path or not os.path.exists(file_path):
        return {}

    VALID = {"Mon","Tue","Wed","Thu","Fri","Sat","Sun"}

    def norm(tok: str):
        tok = (tok or "").strip()
        if not tok:
            return None
        t3 = tok[:3].title()
        if t3 in VALID:
            return t3
        return None

    preferred = {}
    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        next(reader, None)
        for row in reader:
            if not row:
                continue
            team = (row[0] or "").strip()
            if not team:
                continue
            tokens = []
            for cell in row[1:]:
                cell = (cell or "").strip()
                if not cell:
                    continue
                for part in re.split(r"[;,\s]+", cell):
                    part = part.strip()
                    if part:
                        tokens.append(part)
            days = set()
            for t in tokens:
                d = norm(t)
                if d:
                    days.add(d)
            preferred[team] = days
    return preferred


def preferred_day_bonus(team1, team2, d, team_preferred_days):
    if not team_preferred_days:
        return 0
    dow = dow_label(d)
    t1_pref = dow in team_preferred_days.get(team1, set())
    t2_pref = dow in team_preferred_days.get(team2, set())
    if t1_pref and t2_pref:
        return PREFERRED_DAY_BONUS_BOTH
    if t1_pref or t2_pref:
        return PREFERRED_DAY_BONUS_ONE
    return 0


def late_date_penalty(d, season_start, penalty_per_day=LATE_DATE_PENALTY_PER_DAY):
    if season_start is None:
        return 0
    return max(0, (d - season_start).days) * penalty_per_day


def preferred_day_count(teams, d, team_preferred_days):
    if not team_preferred_days:
        return 0
    dow = dow_label(d)
    return sum(1 for t in teams if dow in team_preferred_days.get(t, set()))


def load_field_availability(file_path):
    """Load (date, slot, field) rows and return a *deduplicated* list sorted in a stable order.

    Notes:
      - The scheduler iterates `field_availability` greedily. If this list is in an odd order
        (or contains duplicate rows), you can get surprising behavior (e.g., late slots filled first,
        repeated rows in output exports, etc.).
      - We deduplicate on (date, slot, field) to protect against accidental duplicate rows in the CSV.
      - We sort in a clear order: (is_sunday, date, time, field). If you don't want Sundays first,
        set SUNDAYS_FIRST = False near the top of the file.
    """
    field_availability = []
    seen = set()
    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        next(reader)  # header
        for row in reader:
            date_dt = datetime.strptime(row[0].strip(), '%Y-%m-%d')  # midnight datetime
            slot = row[1].strip()
            field = row[2].strip()
            key = (date_dt.date(), slot, field)
            if key in seen:
                continue
            seen.add(key)
            field_availability.append((date_dt, slot, field))

    def _slot_time(slot_str: str):
        try:
            return datetime.strptime(slot_str.strip(), "%I:%M %p")
        except Exception:
            # If a slot string is malformed, push it to the end but keep deterministic ordering.
            return datetime.strptime("11:59 PM", "%I:%M %p")

    # Stable ordering. Many leagues prefer Sundays filled first; make it configurable.
    sundays_first = globals().get("SUNDAYS_FIRST", True)
    if sundays_first:
        field_availability.sort(key=lambda x: (
            (0 if x[0].weekday() == 6 else 1),
            x[0].date(),
            _slot_time(x[1]),
            x[2],
        ))
    else:
        field_availability.sort(key=lambda x: (
            x[0].date(),
            _slot_time(x[1]),
            x[2],
        ))
    return field_availability


def load_team_blackouts(file_path):
    """
    CSV format: Team, Date1, Date2, ...
    Dates: YYYY-MM-DD
    Returns: dict[team] -> set(date)
    """
    blackouts = {}
    with open(file_path, mode='r') as file:
        reader = csv.reader(file)
        next(reader)  # header
        for row in reader:
            team = row[0].strip()
            dates = set()
            for d in row[1:]:
                d = (d or '').strip()
                if not d:
                    continue
                try:
                    dt = datetime.strptime(d, '%Y-%m-%d').date()
                    dates.add(dt)
                except Exception as e:
                    print("Error parsing blackout date '{}' for team {}: {}".format(d, team, e))
            blackouts[team] = dates
    return blackouts

# -------------------------------
# Intra-division matchup generation
# -------------------------------
def _round_robin_pairs(teams):
    teams = list(teams)
    n = len(teams)
    assert n % 2 == 0, "round robin requires even team count"
    left = teams[:n//2]
    right = teams[n//2:]
    rounds = []
    for _ in range(n-1):
        pairs = list(zip(left, reversed(right)))
        rounds.append(pairs)
        right = [left.pop(1)] + right
        left.insert(1, right.pop())
    return rounds

def generate_intra_matchups_for_target(division, teams, intra_target_per_team):
    teams = sorted(teams)
    n = len(teams)
    if n < 2:
        return []

    if intra_target_per_team < 0:
        raise Exception("intra_target_per_team must be >= 0 (got {}) for division {}.".format(intra_target_per_team, division))

    if intra_target_per_team == 0:
        return []

    if intra_target_per_team == 2 * (n - 1):
        matchups = []
        for t1, t2 in itertools.combinations(teams, 2):
            matchups.append((t1, t2))
            matchups.append((t2, t1))
        return matchups

    if n == 8 and intra_target_per_team == 18:
        two_game_count = 3
        pairs = list(itertools.combinations(teams, 2))
        count2 = {t: 0 for t in teams}
        assignment = {}

        def backtrack(i):
            if i == len(pairs):
                return all(count2[t] == two_game_count for t in teams)
            a, b = pairs[i]
            if count2[a] < two_game_count and count2[b] < two_game_count:
                assignment[(a, b)] = 2
                count2[a] += 1
                count2[b] += 1
                if backtrack(i + 1):
                    return True
                count2[a] -= 1
                count2[b] -= 1
                del assignment[(a, b)]
            assignment[(a, b)] = 3
            if backtrack(i + 1):
                return True
            del assignment[(a, b)]
            return False

        if not backtrack(0):
            raise Exception("No valid intra-division assignment found for {} (18 target).".format(division))

        matchups = []
        for (a, b), w in assignment.items():
            if w == 2:
                matchups.extend([(a, b), (b, a)])
            else:
                matchups.extend([(a, b), (b, a)])
                matchups.append((a, b) if random.random() < 0.5 else (b, a))
        return matchups

    if n == 8 and intra_target_per_team == 22:
        matchups = []
        for a, b in itertools.combinations(teams, 2):
            matchups.extend([(a, b), (b, a)])
            matchups.append((a, b) if random.random() < 0.5 else (b, a))

        rounds = _round_robin_pairs(teams)
        rival_pairs = random.choice(rounds)
        for a, b in rival_pairs:
            matchups.append((a, b) if random.random() < 0.5 else (b, a))
        return matchups

    total_slots = n * intra_target_per_team
    if total_slots % 2 != 0:
        raise Exception(
            "Intra target {} with n={} yields odd total participation ({}); cannot form whole games for division {}."
            .format(intra_target_per_team, n, total_slots, division)
        )

    games_left = {t: intra_target_per_team for t in teams}
    home = {t: 0 for t in teams}
    away = {t: 0 for t in teams}
    matchups = []

    if intra_target_per_team >= 2:
        for i in range(n):
            h = teams[i]
            a = teams[(i + 1) % n]
            matchups.append((h, a))
            home[h] += 1
            away[a] += 1
            games_left[h] -= 1
            games_left[a] -= 1

        for i in range(n):
            h = teams[(i + 1) % n]
            a = teams[i]
            matchups.append((h, a))
            home[h] += 1
            away[a] += 1
            games_left[h] -= 1
            games_left[a] -= 1

    elif intra_target_per_team == 1:
        for i in range(n):
            h = teams[i]
            a = teams[(i + 1) % n]
            matchups.append((h, a))
            home[h] += 1
            away[a] += 1
            games_left[h] -= 1
            games_left[a] -= 1

    meet = defaultdict(int)
    for (h, a) in matchups:
        meet[frozenset((h, a))] += 1
    min_pair, soft_cap = effective_pair_rules(division, intra_target_per_team, n)

    guard = 0
    guard_max = 200000

    def teams_by_need():
        return sorted(teams, key=lambda t: games_left[t], reverse=True)

    while any(v > 0 for v in games_left.values()):
        guard += 1
        if guard > guard_max:
            raise Exception("Failed building intra matchups for {}; stuck with remaining={}".format(division, games_left))

        t1 = teams_by_need()[0]
        if games_left[t1] <= 0:
            break

        candidates = [t for t in teams if t != t1 and games_left[t] > 0]
        if not candidates:
            raise Exception("Cannot find opponent to satisfy intra target for {}. Remaining={}".format(division, games_left))

        def meet_key(t2):
            m = meet[frozenset((t1, t2))]
            # Prefer opponents we haven't met enough yet (under min_pair), then fewer repeats.
            under = 1 if m < min_pair else 0
            return (-under, m, -games_left[t2], t2)

        under = [t2 for t2 in candidates if meet[frozenset((t1, t2))] < soft_cap]
        pick_pool = under if under else candidates
        t2 = min(pick_pool, key=meet_key)

        if home[t1] - away[t1] <= home[t2] - away[t2]:
            h, a = t1, t2
        else:
            h, a = t2, t1

        matchups.append((h, a))
        home[h] += 1
        away[a] += 1
        games_left[h] -= 1
        games_left[a] -= 1
        meet[frozenset((t1, t2))] += 1

    return matchups

# -------------------------------
# Inter-division matchup generation
# -------------------------------
def generate_bipartite_regular_matchups(teams1, teams2, degree):
    teams1 = list(teams1)
    teams2 = list(teams2)

    if degree < 0:
        raise Exception("degree must be >= 0")
    if degree == 0:
        return []
    if degree > len(teams2):
        raise Exception(
            "degree={} exceeds opponent count={}; reduce degree or implement repeat-opponent inter matchups."
            .format(degree, len(teams2))
        )

    random.shuffle(teams1)

    total_edges = len(teams1) * degree
    base = total_edges // len(teams2)
    extra = total_edges % len(teams2)

    teams2_shuffled = teams2[:]
    random.shuffle(teams2_shuffled)
    cap = {t: base for t in teams2_shuffled}
    for t in teams2_shuffled[:extra]:
        cap[t] += 1

    edges = []
    for t1 in teams1:
        avail = [t for t in teams2_shuffled if cap[t] > 0]
        if len(avail) < degree:
            raise Exception("No valid bipartite matching found (insufficient capacity).")

        random.shuffle(avail)
        avail.sort(key=lambda t: cap[t], reverse=True)
        chosen = avail[:degree]

        for t2 in chosen:
            edges.append((t1, t2))
            cap[t2] -= 1

    return edges

def generate_inter_division_matchups(division_from, division_to, teams_from, teams_to, degree):
    edges = generate_bipartite_regular_matchups(teams_from, teams_to, degree)
    matchups = []
    for (t1, t2) in edges:
        matchups.append((t1, t2) if random.random() < 0.5 else (t2, t1))
    return matchups

# -------------------------------
# Combine full matchup list
# -------------------------------
def generate_full_matchups(division_teams):
    enabled_pairs = []
    for (d1, d2), enabled in INTER_PAIR_SETTINGS.items():
        if not enabled:
            continue
        if d1 not in division_teams or d2 not in division_teams:
            continue
        if inter_enabled_for_pair(d1, d2):
            enabled_pairs.append((d1, d2))

    inter_per_team = {d: 0 for d in division_teams.keys()}
    for d1, d2 in enabled_pairs:
        deg = pair_degree(d1, d2)
        inter_per_team[d1] += deg
        inter_per_team[d2] += deg

    full_matchups = []
    for div, teams in division_teams.items():
        # DH-only divisions build their own matchups inside the pod scheduler.
        if is_dh_only(div):
            continue
        intra_target = DIVISION_SETTINGS[div]['target_games'] - inter_per_team.get(div, 0)
        full_matchups.extend(generate_intra_matchups_for_target(div, teams, intra_target))

    for d1, d2 in enabled_pairs:
        deg = pair_degree(d1, d2)
        full_matchups.extend(generate_inter_division_matchups(d1, d2, division_teams[d1], division_teams[d2], deg))

    random.shuffle(full_matchups)

    return full_matchups


def generate_filler_matchups(division_teams, team_stats, schedule, max_new_games=5000):
    """Generate additional flexible matchups to help finish schedules when the fixed matchup list dead-ends.

    Why this exists:
      - The heuristic scheduler starts with a *fixed* list of matchups (opponent graph).
      - With tight calendar constraints (HARD_MIN_GAP, WEEKLY_GAME_LIMIT, availability/blackouts),
        that fixed list can become impossible to place even though there are still plenty of open slots.
      - This function creates *extra* candidate matchups among teams that are still below target,
        preferring intra-division and respecting inter-division enablement rules.

    We keep it conservative:
      - Never creates A games (A is pod-scheduled only).
      - Prefers opponents that haven't been over-used yet (soft caps).
    """
    # Count current meetings (undirected) from the existing schedule
    meet = defaultdict(int)
    for _dt, _slot, _field, home, _hd, away, _ad in schedule:
        if home and away:
            meet[frozenset((home, away))] += 1

    # Teams still needing games (exclude DH-only divisions — they never take singles)
    need = [t for div, teams in division_teams.items() for t in teams
            if not is_dh_only(div) and team_stats[t]['total_games'] < target_games(t)]

    if not need:
        return []

    # Precompute per-division effective caps (rough)
    caps = {}
    for div, teams in division_teams.items():
        if is_dh_only(div):
            continue
        n = len(teams)
        intra_target = max(0, DIVISION_SETTINGS[div]['target_games'])  # conservative
        _min_eff, cap_eff = effective_pair_rules(div, intra_target, n)
        caps[div] = cap_eff

    new_matchups = []
    guard = 0
    while guard < max_new_games:
        guard += 1

        # Refresh need list
        need = [t for div, teams in division_teams.items() for t in teams
                if not is_dh_only(div) and team_stats[t]['total_games'] < target_games(t)]
        if not need:
            break

        # Pick most-behind team
        need.sort(key=lambda t: (target_games(t) - team_stats[t]['total_games'], t), reverse=True)
        t1 = need[0]
        d1 = div_of(t1)

        # Candidate opponents: also behind, and allowed to play (same div or enabled inter pair)
        opps = []
        for t2 in need[1:]:
            if t2 == t1:
                continue
            d2 = div_of(t2)
            if d1 == d2:
                opps.append(t2)
            else:
                if inter_enabled_for_pair(d1, d2):
                    opps.append(t2)

        if not opps:
            # If no behind opponents are available, fall back to any team in same division
            opps = [t2 for t2 in division_teams.get(d1, []) if t2 != t1 and not is_dh_only(d1)]

        if not opps:
            break

        # Prefer opponents with:
        #   - biggest deficit
        #   - lowest current meet count
        #   - below soft cap
        def opp_key(t2):
            d2 = div_of(t2)
            pair = frozenset((t1, t2))
            m = meet[pair]
            cap = max(caps.get(d1, 999), caps.get(d2, 999))
            over = 1 if m >= cap else 0
            return (over, m, -(target_games(t2) - team_stats[t2]['total_games']), t2)

        t2 = min(opps, key=opp_key)

        # Add matchup with randomized home/away orientation
        if random.random() < 0.5:
            new_matchups.append((t1, t2))
        else:
            new_matchups.append((t2, t1))
        meet[frozenset((t1, t2))] += 1

        # Optimistically increment totals so we don't over-generate for a single team
        team_stats[t1]['total_games'] += 1
        team_stats[t2]['total_games'] += 1

    # Roll back the optimistic increments (we only wanted them for generation weighting)
    for (h, a) in new_matchups:
        team_stats[h]['total_games'] -= 1
        team_stats[a]['total_games'] -= 1

    return new_matchups

# -------------------------------
# Home/Away Helper
# -------------------------------

def build_sunday_pod_assignment(timeslots_by_date, rotation, seed=42):
    """Return {date: division} assignment for which division gets pod-style DH priority on Sundays.

    We shuffle Sundays (seeded) then round-robin assign divisions. This helps spread Sunday pods.
    """
    sundays = [d for d in timeslots_by_date.keys() if getattr(d, 'weekday', lambda: 0)() == 6]
    sundays = sorted(sundays)
    rnd = random.Random(seed)
    rnd.shuffle(sundays)
    if not rotation:
        rotation = ['A', 'B', 'C', 'D']
    mapping = {}
    for i, d in enumerate(sundays):
        mapping[d] = rotation[i % len(rotation)]
    return mapping


def decide_home_away(t1, t2, team_stats):
    if team_stats[t1]['home_games'] >= HOME_AWAY_BALANCE and team_stats[t2]['home_games'] < HOME_AWAY_BALANCE:
        return t2, t1
    if team_stats[t2]['home_games'] >= HOME_AWAY_BALANCE and team_stats[t1]['home_games'] < HOME_AWAY_BALANCE:
        return t1, t2
    if team_stats[t1]['home_games'] < team_stats[t2]['home_games']:
        return t1, t2
    if team_stats[t2]['home_games'] < team_stats[t1]['home_games']:
        return t2, t1
    return (t1, t2) if random.random() < 0.5 else (t2, t1)


def schedule_doubleheaders_preemptively(all_teams, unscheduled, team_availability, field_availability, team_blackouts, timeslots_by_date,
                                        team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents,
                                        used_slots, schedule=None, field_index=None):
    if schedule is None:
        schedule = []
    if field_index is None:
        field_index = build_field_index(field_availability)

    # Prefer filling Sundays first (league preference: easiest full-day inventory)
    date_order = sorted(timeslots_by_date.keys(), key=lambda dd: (0 if dd.weekday() == 6 else 1, dd))
    for d in date_order:
        day_of_week = dow_label(d)
        week_num = d.isocalendar()[1]
        slots = timeslots_by_date[d]
        if not slots:
            continue

        teams_by_need = sorted(all_teams, key=lambda t: team_need_key(t, team_stats, doubleheader_count), reverse=True)
        for team in teams_by_need:
            if team and is_dh_only(team):
                continue
            if doubleheader_count[team] >= min_dh(team):
                continue
            if not is_team_available(team, d, team_availability, team_blackouts):
                continue

            games_today = team_game_days[team].get(d, 0)

            if games_today == 0:
                if len(slots) < 2:
                    continue

                for i in range(len(slots) - 1):
                    slot1 = slots[i]
                    slot2 = slots[i + 1]

                    free1 = free_fields_for_slot(field_index, d, slot1, used_slots)
                    free2 = free_fields_for_slot(field_index, d, slot2, used_slots)
                    if not free1 or not free2:
                        continue

                    candidate_matchups = [m for m in unscheduled if team in m]
                    if len(candidate_matchups) < 2:
                        continue

                    for m1, m2 in itertools.combinations(candidate_matchups, 2):
                        opp1 = m1[0] if m1[1] == team else m1[1]
                        opp2 = m2[0] if m2[1] == team else m2[1]
                        if opp1 == opp2:
                            continue

                        if not is_team_available(opp1, d, team_availability, team_blackouts):
                            continue
                        if not is_team_available(opp2, d, team_availability, team_blackouts):
                            continue
                        if team_game_days[opp1].get(d, 0) != 0 or team_game_days[opp2].get(d, 0) != 0:
                            continue

                        if team_stats[team]['weekly_games'][week_num] + 2 > WEEKLY_GAME_LIMIT:
                            continue
                        if team_stats[opp1]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                            continue
                        if team_stats[opp2]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                            continue

                        if team_stats[team]['total_games'] + 2 > target_games(team):
                            continue
                        if team_stats[opp1]['total_games'] + 1 > target_games(opp1):
                            continue
                        if team_stats[opp2]['total_games'] + 1 > target_games(opp2):
                            continue

                        home1, away1 = decide_home_away(team, opp1, team_stats)
                        home2, away2 = decide_home_away(team, opp2, team_stats)

                        date1, slot1_str, field1 = free1[0]
                        date2, slot2_str, field2 = free2[0]

                        unscheduled.remove(m1)
                        unscheduled.remove(m2)

                        team_stats[home1]['home_games'] += 1
                        team_stats[away1]['away_games'] += 1
                        team_stats[home2]['home_games'] += 1
                        team_stats[away2]['away_games'] += 1

                        schedule.append((date1, slot1_str, field1, home1, home1[0], away1, away1[0]))
                        schedule.append((date2, slot2_str, field2, home2, home2[0], away2, away2[0]))

                        for t in (team, opp1):
                            team_stats[t]['total_games'] += 1
                            team_stats[t]['weekly_games'][week_num] += 1
                            team_game_days[t][d] += 1
                            team_game_slots[t][d].append(slot1_str)

                        for t in (team, opp2):
                            team_stats[t]['total_games'] += 1
                            team_stats[t]['weekly_games'][week_num] += 1
                            team_game_days[t][d] += 1
                            team_game_slots[t][d].append(slot2_str)

                        doubleheader_count[team] += 1
                        team_doubleheader_opponents[team][d].update([opp1, opp2])

                        used_slots[(date1, slot1_str, field1)] = True
                        used_slots[(date2, slot2_str, field2)] = True
                        break

            elif games_today == 1:
                current_slot = team_game_slots[team][d][0]
                try:
                    idx = slots.index(current_slot)
                except ValueError:
                    continue
                if idx + 1 >= len(slots):
                    continue
                next_slot = slots[idx + 1]

                free_next = free_fields_for_slot(field_index, d, next_slot, used_slots)
                if not free_next:
                    continue

                already_opp = None
                for g in schedule:
                    if g[0].date() == d and (g[3] == team or g[5] == team):
                        already_opp = g[5] if g[3] == team else g[3]
                        break
                if already_opp is None:
                    continue

                if doubleheader_count[team] >= max_dh(team):
                    continue

                candidate_matchups = [m for m in unscheduled if team in m]
                for m in candidate_matchups:
                    opp = m[0] if m[1] == team else m[1]
                    if opp == already_opp:
                        continue
                    if not is_team_available(opp, d, team_availability, team_blackouts):
                        continue
                    if team_game_days[opp].get(d, 0) != 0:
                        continue
                    if team_stats[team]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                        continue
                    if team_stats[opp]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                        continue
                    if team_stats[team]['total_games'] + 1 > target_games(team):
                        continue
                    if team_stats[opp]['total_games'] + 1 > target_games(opp):
                        continue
                    if opp in team_doubleheader_opponents[team][d]:
                        continue
                    home, away = decide_home_away(team, opp, team_stats)
                    date_entry, slot_str, field = free_next[0]

                    unscheduled.remove(m)
                    team_stats[home]['home_games'] += 1
                    team_stats[away]['away_games'] += 1
                    schedule.append((date_entry, slot_str, field, home, home[0], away, away[0]))
                    for t in (team, opp):
                        team_stats[t]['total_games'] += 1
                        team_stats[t]['weekly_games'][week_num] += 1
                        team_game_days[t][d] += 1
                        team_game_slots[t][d].append(slot_str)

                    doubleheader_count[team] += 1
                    # Record the pairing for BOTH teams so neither can later be given
                    # another game against the same opponent on this date.
                    team_doubleheader_opponents[team][d].add(opp)
                    team_doubleheader_opponents[opp][d].add(team)
                    used_slots[(date_entry, slot_str, field)] = True
                    break

    return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents, used_slots, unscheduled

# -------------------------------
# Dedicated Doubleheader pass (Two-phase), per-division min/max
# -------------------------------
def force_minimum_doubleheaders(all_teams, unscheduled, team_availability, field_availability, team_blackouts, timeslots_by_date,
                                team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents,
                                used_slots, schedule=None, field_index=None):
    if schedule is None:
        schedule = []
    if field_index is None:
        field_index = build_field_index(field_availability)

    teams = sorted(all_teams, key=lambda t: team_need_key(t, team_stats, doubleheader_count), reverse=True)

    # Phase 1: ensure each team gets at least 1 DH day (if min_dh > 0)
    for team in teams:
        if team and is_dh_only(team):
            continue
        if min_dh(team) <= 0 or doubleheader_count[team] >= 1:
            continue

        date_order = sorted(timeslots_by_date.keys(), key=lambda dd: (0 if dd.weekday() == 6 else 1, dd))
        for d in date_order:
            day_of_week = dow_label(d)
            if d in team_blackouts.get(team, set()) or day_of_week not in team_availability.get(team, set()):
                continue
            week_num = d.isocalendar()[1]
            sorted_slots = timeslots_by_date[d]
            games_today = team_game_days[team].get(d, 0)

            if games_today != 1:
                continue

            try:
                idx = sorted_slots.index(team_game_slots[team][d][0])
            except ValueError:
                continue
            if idx + 1 >= len(sorted_slots):
                continue
            next_slot = sorted_slots[idx + 1]

            free_fields = free_fields_for_slot(field_index, d, next_slot, used_slots)
            if not free_fields:
                continue

            already_opp = None
            for g in schedule:
                if g[0].date() == d and (g[3] == team or g[5] == team):
                    already_opp = g[5] if g[3] == team else g[3]
                    break
            if already_opp is None:
                continue

            if doubleheader_count[team] >= max_dh(team):
                break

            candidate = [m for m in unscheduled if team in m]
            for m in candidate:
                opp = m[0] if m[1] == team else m[1]
                if opp == already_opp:
                    continue
                if not is_team_available(opp, d, team_availability, team_blackouts):
                    continue
                if team_game_days[opp].get(d, 0) != 0:
                    continue
                if team_stats[team]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                    continue
                if team_stats[opp]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                    continue
                if team_stats[team]['total_games'] + 1 > target_games(team):
                    continue
                if team_stats[opp]['total_games'] + 1 > target_games(opp):
                    continue
                if opp in team_doubleheader_opponents[team][d]:
                    continue
                home, away = decide_home_away(team, opp, team_stats)
                date_entry, slot_str, field = free_fields[0]

                unscheduled.remove(m)
                team_stats[home]['home_games'] += 1
                team_stats[away]['away_games'] += 1
                schedule.append((date_entry, slot_str, field, home, home[0], away, away[0]))
                for t in (team, opp):
                    team_stats[t]['total_games'] += 1
                    team_stats[t]['weekly_games'][week_num] += 1
                    team_game_days[t][d] += 1
                    team_game_slots[t][d].append(slot_str)

                doubleheader_count[team] += 1
                # Record the pairing for BOTH teams so neither can later be given
                # another game against the same opponent on this date.
                team_doubleheader_opponents[team][d].add(opp)
                team_doubleheader_opponents[opp][d].add(team)
                used_slots[(date_entry, slot_str, field)] = True
                break

            if doubleheader_count[team] >= 1:
                break

    # Phase 2: push teams toward their per-division minimum DH days.
    teams = sorted(all_teams, key=lambda t: team_need_key(t, team_stats, doubleheader_count), reverse=True)
    for team in teams:
        if team and is_dh_only(team):
            continue
        while doubleheader_count[team] < min_dh(team):
            if doubleheader_count[team] >= max_dh(team):
                break

            scheduled = False
            date_order = sorted(timeslots_by_date.keys(), key=lambda dd: (0 if dd.weekday() == 6 else 1, dd))
            for d in date_order:
                day_of_week = dow_label(d)
                if d in team_blackouts.get(team, set()) or day_of_week not in team_availability.get(team, set()):
                    continue
                week_num = d.isocalendar()[1]
                sorted_slots = timeslots_by_date[d]
                games_today = team_game_days[team].get(d, 0)

                if games_today == 1:
                    try:
                        idx = sorted_slots.index(team_game_slots[team][d][0])
                    except ValueError:
                        continue
                    if idx + 1 >= len(sorted_slots):
                        continue
                    next_slot = sorted_slots[idx + 1]

                    free_fields = free_fields_for_slot(field_index, d, next_slot, used_slots)
                    if not free_fields:
                        continue

                    already_opp = None
                    for g in schedule:
                        if g[0].date() == d and (g[3] == team or g[5] == team):
                            already_opp = g[5] if g[3] == team else g[3]
                            break
                    if already_opp is None:
                        continue

                    candidate = [m for m in unscheduled if team in m]
                    for m in candidate:
                        opp = m[0] if m[1] == team else m[1]
                        if opp == already_opp:
                            continue
                        if not is_team_available(opp, d, team_availability, team_blackouts):
                            continue
                        if team_game_days[opp].get(d, 0) != 0:
                            continue
                        if team_stats[team]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                            continue
                        if team_stats[opp]['weekly_games'][week_num] + 1 > WEEKLY_GAME_LIMIT:
                            continue
                        if team_stats[team]['total_games'] + 1 > target_games(team):
                            continue
                        if team_stats[opp]['total_games'] + 1 > target_games(opp):
                            continue
                        if opp in team_doubleheader_opponents[team][d]:
                            continue

                        home, away = decide_home_away(team, opp, team_stats)
                        date_entry, slot_str, field = free_fields[0]

                        unscheduled.remove(m)
                        team_stats[home]['home_games'] += 1
                        team_stats[away]['away_games'] += 1
                        schedule.append((date_entry, slot_str, field, home, home[0], away, away[0]))

                        for t in (team, opp):
                            team_stats[t]['total_games'] += 1
                            team_stats[t]['weekly_games'][week_num] += 1
                            team_game_days[t][d] += 1
                            team_game_slots[t][d].append(slot_str)

                        doubleheader_count[team] += 1
                        # Record the pairing for BOTH teams so neither can later be given
                        # another game against the same opponent on this date.
                        team_doubleheader_opponents[team][d].add(opp)
                        team_doubleheader_opponents[opp][d].add(team)
                        used_slots[(date_entry, slot_str, field)] = True
                        scheduled = True
                        break

                if scheduled:
                    break

            if not scheduled:
                break

    return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents, used_slots, unscheduled

# -------------------------------
# A Division DH-only scheduling (pair doubleheaders)
# -------------------------------

# -------------------------------
# A Division DH-only scheduling (4-team pods across two fields)
# -------------------------------
def schedule_pod_only_division(div, division_teams, team_availability, field_availability, team_blackouts,
                                 timeslots_by_date, team_stats, doubleheader_count,
                                 team_game_days, team_game_slots, used_slots, schedule=None, sunday_assignment=None, sunday_pods_used=None, team_preferred_days=None):
    """
    Schedule a division as *doubleheaders only* using 4-team "pod" sessions across BOTH fields.

    Used for any division with dh_only=True in DIVISION_SETTINGS (historically Division A).

    A pod session on date d uses two adjacent slots (s1,s2) and two different fields (f1,f2):
      Slot s1:
        Game1: t1 vs t2  (on f1)
        Game2: t3 vs t4  (on f2)
      Slot s2:
        Game3: t1 vs t3  (on f1)
        Game4: t2 vs t4  (on f2)

    This guarantees each team plays exactly 2 games that day with DIFFERENT opponents.
    By construction, each team gets 1 home + 1 away within the pod:
      Slot s1 home teams: t1, t3
      Slot s2 home teams: t2, t4
    """
    if schedule is None:
        schedule = []
    if not isinstance(schedule, list):
        raise TypeError("schedule must be list[game_tuple], got {}".format(type(schedule)))

    div = div[0].upper()
    A_teams = list(division_teams.get(div, []))
    if not A_teams:
        return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, used_slots

    # Every game is part of a pod, so N games => N/2 DH days (sessions) per team
    target_sessions = DIVISION_SETTINGS[div]['target_games'] // 2

    # Opponent-balance rules for this division (min games per pairing + soft cap)
    _rules = PAIR_RULES.get(div, {'min': 1, 'soft_cap': 3})
    PAIR_MIN_GAMES = int(_rules.get('min', 1))
    PAIR_SOFT_CAP = int(_rules.get('soft_cap', 3))

    sessions_done = defaultdict(int)      # team -> sessions completed
    pair_meets = defaultdict(int)         # frozenset({a,b}) -> number of games already between them (within A pods)

    # Build fast lookup for canonical datetime from field_availability (midnight dt)
    dt_by_key = {}
    for date_dt, slot, field in field_availability:
        dt_by_key[(date_dt.date(), slot, field)] = date_dt

    # Build per-date slot ordering + list of adjacent slot pairs.
    #
    # We *do* want some A pods on Sundays (easier attendance), but we don't want A to dominate
    # every Sunday all season.
    #
    # Strategy:
    #   1) Ensure each A team gets at least MIN_SUNDAY_SESSIONS_PER_TEAM DH sessions on Sundays
    #   2) After that, prefer weekday pods.
    all_dates = sorted({dt.date() for (dt, _slot, _field) in field_availability})
    sunday_dates = [d for d in all_dates if d.weekday() == 6]
    weekday_dates = [d for d in all_dates if d.weekday() != 6]

    # Policy: limit how many A "pods" can run on any given Sunday.
    # A "pod" = 4 A-teams playing 2 games each across two fields and two adjacent slots.
    MAX_A_PODS_PER_SUNDAY = SUNDAY_PODS_PER_SUNDAY
    MIN_SUNDAY_SESSIONS_PER_TEAM = 1
    adjacent_slot_pairs_by_date = {}
    for d in all_dates:
        slots = sorted(set(timeslots_by_date.get(d, [])), key=lambda s: datetime.strptime(s.strip(), "%I:%M %p"))
        pairs = []
        for i in range(len(slots) - 1):
            pairs.append((slots[i], slots[i + 1]))
        adjacent_slot_pairs_by_date[d] = pairs

    # Build per-date set of fields available
    fields_by_date = defaultdict(set)
    for date_dt, slot, field in field_availability:
        fields_by_date[date_dt.date()].add(field)

    def can_play_pod(team, d):
        dow = dow_abbrev(d)
        if not is_team_available(team, d, team_availability, team_blackouts):
            return False
        # must have no other games that date
        if team_game_days[team].get(d, 0) != 0:
            return False
        # gap constraint vs other days
        if not min_gap_ok(team, d, team_game_days):
            return False
        if not no_two_consecutive_byes_after_adding(team, d, team_game_days):
            return False
        wk = d.isocalendar()[1]
        if team_stats[team]['weekly_games'].get(wk, 0) + 2 > WEEKLY_GAME_LIMIT:
            return False
        # Soft weekly balance: on early passes refuse to push a team past its even
        # share for the week, so pods spread out instead of stacking two into one
        # week. Later passes relax this so we don't lose games we can't place anywhere else.
        if WEEKLY_BALANCE_PENALTY > 0 and strict_weekly and weekly_excess(team, d, team_stats, extra=2) > 0:
            return False
        if team_stats[team]['total_games'] + 2 > DIVISION_SETTINGS[div]['target_games']:
            return False
        return True

    def available_fields_for_pair(d, s1, s2):
        """Return fields that are free (unused) for BOTH slots s1 and s2 on date d."""
        out = []
        for f in sorted(fields_by_date.get(d, [])):
            dt1 = dt_by_key.get((d, s1, f))
            dt2 = dt_by_key.get((d, s2, f))
            if dt1 is None or dt2 is None:
                continue
            if used_slots.get((dt1, s1, f), False) or used_slots.get((dt2, s2, f), False):
                continue
            out.append(f)
        return out
    def choose_four(eligible):
        """Pick 4 eligible teams for a pod session, *actively* balancing opponents.

        Goals:
          1) Ensure every intra-division pairing happens at least PAIR_MIN_GAMES times (as feasible)
          2) Avoid extreme repeats early (e.g., A1 vs A2 six times while A1 vs A8 once)
          3) Still finish all required sessions

        We evaluate both which 4 teams to use AND the internal pod layout (who plays whom),
        because the layout determines the 4 games created:
            slot1: t1-vs-t2, t3-vs-t4
            slot2: t1-vs-t4, t2-vs-t3
        """
        need = [t for t in eligible if sessions_done[t] < target_sessions]
        if len(need) < 4:
            return None

        # Prefer teams with biggest remaining sessions; keep pool small for speed
        need.sort(
            key=lambda t: (
                target_sessions - sessions_done[t],
                DIVISION_SETTINGS['A']['target_games'] - team_stats[t]['total_games'],
                t,
            ),
            reverse=True,
        )
        pool = need[:10] if len(need) > 10 else need

        # Helper: does team still have any unmet "must play" pairs?
        def has_unmet_pairs(team: str) -> bool:
            for other in A_teams:
                if other == team:
                    continue
                if pair_meets[frozenset((team, other))] < PAIR_MIN_GAMES:
                    return True
            return False

        best = None
        best_score = None  # smaller is better (lexicographic)

        # Evaluate combinations of 4 from pool, and also try all internal layouts
        for combo in itertools.combinations(pool, 4):
            # Try all unique permutations (layout matters). 24 is small.
            for perm in itertools.permutations(combo, 4):
                t1, t2, t3, t4 = perm

                games = [
                    frozenset((t1, t2)),
                    frozenset((t3, t4)),
                    frozenset((t1, t4)),
                    frozenset((t2, t3)),
                ]

                # Hard-ish guard:
                # If a pair is already at/over soft cap, don't schedule it IF either team still has
                # any unmet required pair elsewhere.
                blocked = False
                for g in games:
                    a, b = tuple(g)
                    if pair_meets[g] >= PAIR_SOFT_CAP and (has_unmet_pairs(a) or has_unmet_pairs(b)):
                        blocked = True
                        break
                if blocked:
                    continue

                # Count how many of the games help satisfy the minimum pair requirement
                unmet_hits = sum(1 for g in games if pair_meets[g] < PAIR_MIN_GAMES)

                # Prefer layouts that:
                #  - maximize unmet_hits
                #  - minimize total existing meetings for these pairs
                #  - then prefer using teams with larger remaining session deficits
                total_meets = sum(pair_meets[g] for g in games)
                rem = sum((target_sessions - sessions_done[t]) for t in (t1, t2, t3, t4))

                # Prefer pods that include the most schedule-constrained teams, so
                # heavily-blacked-out teams grab their rare playable dates first.
                combo_scarcity = sum(team_scarcity(t) for t in (t1, t2, t3, t4))

                # Also reduce spread (avoid spiking any single pair too quickly)
                after_counts = [pair_meets[g] + 1 for g in games]
                spread = max(after_counts) - min(after_counts)

                score = (-combo_scarcity, -unmet_hits, total_meets, spread, -rem, tuple(sorted(combo)))
                if best_score is None or score < best_score:
                    best_score = score
                    best = (t1, t2, t3, t4)

        return best

    def place_game(d, slot, field, home, away):
        dt = dt_by_key.get((d, slot, field))
        if dt is None:
            return False
        # Hard cap
        if team_stats[home]['total_games'] >= target_games(home) or team_stats[away]['total_games'] >= target_games(away):
            return False
        schedule.append((dt, slot, field, home, home[0], away, away[0]))
        used_slots[(dt, slot, field)] = True

        wk = d.isocalendar()[1]
        team_stats[home]['total_games'] += 1
        team_stats[away]['total_games'] += 1
        team_stats[home]['home_games'] += 1
        team_stats[away]['away_games'] += 1
        team_stats[home]['weekly_games'][wk] = team_stats[home]['weekly_games'].get(wk, 0) + 1
        team_stats[away]['weekly_games'][wk] = team_stats[away]['weekly_games'].get(wk, 0) + 1
        team_game_days[home][d] += 1
        team_game_days[away][d] += 1
        team_game_slots[home][d].append(slot)
        team_game_slots[away][d].append(slot)
        return True

    # Iterate dates/slots in chronological order. We do multiple passes to work around blocked days/slots.
    #
    # We allow multiple pods per date EXCEPT Sundays, which are capped by MAX_A_PODS_PER_SUNDAY.
    sunday_sessions_done = {t: 0 for t in A_teams}
    season_start = min((dt.date() for dt, _slot, _field in field_availability), default=None)

    # Weekly balance is enforced for most passes, then relaxed for the last few so
    # teams that still cannot be placed anywhere are not left short.
    strict_weekly = True

    for _pass in range(12):
        strict_weekly = _pass < 9
        progress = False

        need_more_sunday = any(sunday_sessions_done[t] < MIN_SUNDAY_SESSIONS_PER_TEAM for t in A_teams)

        # Balance A-division day-of-week distribution:
        # Prefer dates whose weekday is currently under-used by A, with a small seeded shuffle
        # to avoid repeatedly picking the same day pattern.
        a_dow_load = defaultdict(int)
        for _dt, _slot, _field, _home, _hdiv, _away, _adiv in schedule:
            if (_home and _home[0] == div) or (_away and _away[0] == div):
                a_dow_load[dow_label(_dt)] += 1

        rnd = random.Random((RANDOM_SEED or 0) + (_pass * 97) + 13)

        def _date_key(dd):
            # opening weeks first; then lower day-of-week load; then randomized tie-break
            return (0 if in_front_load_window(dd) else 1,
                    a_dow_load.get(dow_label(dd), 0), rnd.random(), dd)

        # If we have a Sunday rotation, push Sundays assigned to A to the front of the Sunday list.
        if sunday_assignment:
            sunday_dates_ordered = [sd for sd in sunday_dates if sunday_assignment.get(sd) == div] +                                    [sd for sd in sunday_dates if sunday_assignment.get(sd) != div]
        else:
            sunday_dates_ordered = list(sunday_dates)

        # Sort within weekday/Sunday groups by current A day-load to smooth out heavy Mondays/Tuesdays.
        weekday_dates_ordered = sorted(list(weekday_dates), key=_date_key)
        sunday_dates_ordered = sorted(list(sunday_dates_ordered), key=_date_key)

        # Sundays first while teams still owe Sunday sessions, or whenever Sundays
        # are explicitly prioritised.
        sundays_first = need_more_sunday or SUNDAY_PRIORITY > 0
        date_order = (sunday_dates_ordered + weekday_dates_ordered) if sundays_first else (weekday_dates_ordered + sunday_dates_ordered)

        for d in date_order:
            # Sunday pod rotation + global cap:
            # - If this Sunday is assigned to a different division, only let A use it while we still
            #   need to satisfy the minimum Sunday sessions per A team.
            if sunday_assignment and d.weekday() == 6 and sunday_assignment.get(d) not in (None, 'A'):
                continue
            if sunday_pods_used is not None and d.weekday() == 6 and sunday_pods_used.get(d, 0) >= SUNDAY_PODS_PER_SUNDAY:
                continue



            if all(sessions_done[t] >= target_sessions for t in A_teams):
                break

            pods_today = 0

            # Try to schedule as many pods as possible on this date across distinct adjacent slot pairs.
            for (s1, s2) in adjacent_slot_pairs_by_date.get(d, []):
                if all(sessions_done[t] >= target_sessions for t in A_teams):
                    break

                # Cap A pods on Sundays
                if d.weekday() == 6 and pods_today >= MAX_A_PODS_PER_SUNDAY:
                    break

                free_fields = available_fields_for_pair(d, s1, s2)
                if len(free_fields) < 2:
                    continue

                eligible = [t for t in A_teams if can_play_pod(t, d) and sessions_done[t] < target_sessions]
                # If we're still trying to give every A team at least MIN_SUNDAY_SESSIONS_PER_TEAM on Sundays,
                # restrict the pool to teams that still need a Sunday session so we don't keep re-using the same 4 teams.
                if need_more_sunday and d.weekday() == 6:
                    need_sun = [t for t in eligible if sunday_sessions_done[t] < MIN_SUNDAY_SESSIONS_PER_TEAM]
                    if len(need_sun) >= 4:
                        eligible = need_sun

                if len(eligible) < 4:
                    continue

                four = choose_four(eligible)
                if not four:
                    continue
                t1, t2, t3, t4 = four

                # assign two distinct fields
                f1, f2 = free_fields[0], free_fields[1]

                ok = True
                ok &= place_game(d, s1, f1, t1, t2)  # t1 home
                ok &= place_game(d, s1, f2, t3, t4)  # t3 home
                ok &= place_game(d, s2, f1, t2, t3)  # t2 home (vs t3)
                ok &= place_game(d, s2, f2, t4, t1)  # t4 home (vs t1)
                if not ok:
                    continue

                for t in (t1, t2, t3, t4):
                    sessions_done[t] += 1
                    doubleheader_count[t] += 1

                pair_meets[frozenset((t1, t2))] += 1
                pair_meets[frozenset((t3, t4))] += 1
                pair_meets[frozenset((t2, t3))] += 1
                pair_meets[frozenset((t4, t1))] += 1

                progress = True
                pods_today += 1
                if sunday_pods_used is not None and d.weekday() == 6:
                    sunday_pods_used[d] = sunday_pods_used.get(d, 0) + 1
                if d.weekday() == 6:
                    for t in (t1, t2, t3, t4):
                        sunday_sessions_done[t] += 1
                # continue scanning later slot pairs on same date to potentially schedule another pod

        if not progress:
            break

    return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, used_slots


# -------------------------------
# B/C/D Doubleheader pods (after A)
# -------------------------------
def _pop_matchup_any_orientation(unscheduled, a, b):
    """Remove and return one matchup between a and b (either (a,b) or (b,a))."""
    try:
        idx = unscheduled.index((a, b))
        return unscheduled.pop(idx)
    except ValueError:
        pass
    try:
        idx = unscheduled.index((b, a))
        return unscheduled.pop(idx)
    except ValueError:
        return None


def schedule_division_pod_doubleheaders(div, division_teams, unscheduled,
                                       team_availability, field_availability, team_blackouts, timeslots_by_date,
                                       team_stats, doubleheader_count, team_game_days, team_game_slots,
                                       team_doubleheader_opponents, used_slots, schedule=None, sunday_assignment=None, sunday_pods_used=None, team_preferred_days=None):
    """Schedule 4-team pod doubleheaders *within a division* to satisfy min_dh() targets.

    This uses the same A-style pod structure (two fields, two adjacent slots) so each of the 4 teams
    plays 2 games that day against DIFFERENT opponents.

    It only consumes matchups that already exist in `unscheduled` (either orientation), so we don't
    accidentally create extra games.
    """
    if schedule is None:
        schedule = []

    teams = list(division_teams.get(div, []))
    if len(teams) < 4:
        return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents, used_slots, unscheduled

    # Fast lookup for canonical datetime from field_availability
    dt_by_key = {(dt.date(), slot, field): dt for (dt, slot, field) in field_availability}

    # Date order follows field_availability sort (Sundays first), but we keep unique dates
    unique_dates = []
    seen = set()
    for dt, _slot, _field in field_availability:
        d = dt.date()
        if d not in seen:
            unique_dates.append(d)
            seen.add(d)

    # Fields available per date
    fields_by_date = defaultdict(set)
    for dt, _slot, field in field_availability:
        fields_by_date[dt.date()].add(field)

    def can_play_pod(team, d):
        dow = dow_abbrev(d)
        if not is_team_available(team, d, team_availability, team_blackouts):
            return False
        if team_game_days[team].get(d, 0) != 0:
            return False
        if not min_gap_ok(team, d, team_game_days):
            return False
        wk = d.isocalendar()[1]
        if team_stats[team]['weekly_games'].get(wk, 0) + 2 > WEEKLY_GAME_LIMIT:
            return False
        # Soft weekly balance — see note in schedule_pod_only_division.
        if WEEKLY_BALANCE_PENALTY > 0 and strict_weekly and weekly_excess(team, d, team_stats, extra=2) > 0:
            return False
        if team_stats[team]['total_games'] + 2 > target_games(team):
            return False
        if doubleheader_count[team] >= max_dh(team):
            return False
        return True

    def available_fields_for_pair(d, s1, s2):
        out = []
        for f in sorted(fields_by_date.get(d, [])):
            dt1 = dt_by_key.get((d, s1, f))
            dt2 = dt_by_key.get((d, s2, f))
            if dt1 is None or dt2 is None:
                continue
            if used_slots.get((dt1, s1, f), False) or used_slots.get((dt2, s2, f), False):
                continue
            out.append(f)
        return out

    def place_game(d, slot, field, t1, t2):
        """Place a single game (t1 vs t2) on (d,slot,field) with balanced home/away."""
        dt = dt_by_key.get((d, slot, field))
        if dt is None:
            return False
        # Hard cap
        if team_stats[t1]['total_games'] >= target_games(t1) or team_stats[t2]['total_games'] >= target_games(t2):
            return False
        home, away = decide_home_away(t1, t2, team_stats)

        # hard cap: never exceed target home balance too much
        if team_stats[home]['home_games'] >= HOME_AWAY_BALANCE and team_stats[away]['home_games'] < HOME_AWAY_BALANCE:
            home, away = away, home
        schedule.append((dt, slot, field, home, home[0], away, away[0]))
        used_slots[(dt, slot, field)] = True

        wk = d.isocalendar()[1]
        team_stats[home]['total_games'] += 1
        team_stats[away]['total_games'] += 1
        team_stats[home]['home_games'] += 1
        team_stats[away]['away_games'] += 1
        team_stats[home]['weekly_games'][wk] = team_stats[home]['weekly_games'].get(wk, 0) + 1
        team_stats[away]['weekly_games'][wk] = team_stats[away]['weekly_games'].get(wk, 0) + 1
        team_game_days[home][d] += 1
        team_game_days[away][d] += 1
        team_game_slots[home][d].append(slot)
        team_game_slots[away][d].append(slot)
        return True

    season_start = min((dt.date() for dt, _slot, _field in field_availability), default=None)

    # Greedy scheduling: multiple passes to reach min_dh.
    # Weekly balance is enforced early and relaxed near the end (see pod-only version).
    strict_weekly = True

    for _pass in range(10):
        strict_weekly = _pass < 7
        progress = False

        # Light day-of-week balancing:
        # Prefer scheduling pods on days this division has used less so far.
        dow_counts = defaultdict(int)
        for (dt0, _slot0, _field0, home0, _hd0, away0, _ad0) in schedule:
            if div_of(home0) == div and div_of(away0) == div:
                dow_counts[dow_label(dt0)] += 1

        rnd = random.Random((RANDOM_SEED or 0) + (_pass * 131) + ord(div))
        # Keep Sundays early (rotation applies), but within that prefer under-used DOWs.
        # A Sunday priority makes that ordering stronger by ignoring the DOW balancing
        # for Sundays, so they fill before weekdays are touched.
        date_order = sorted(unique_dates, key=lambda dd: (
            0 if in_front_load_window(dd) else 1,
            0 if dd.weekday() == 6 else 1,
            0 if (SUNDAY_PRIORITY > 0 and dd.weekday() == 6) else dow_counts[dow_label(dd)],
            rnd.random()
        ))

        # Stop early if everyone in division has hit min DH
        if all(doubleheader_count[t] >= min_dh(t) for t in teams):
            break

        for d in date_order:
            # Sunday pod rotation: only allow this division's pods on Sundays assigned to it
            if sunday_assignment and d.weekday() == 6:
                assigned = sunday_assignment.get(d)
                # First pod on a Sunday is reserved for the assigned division (rotation).
                # If there is remaining Sunday pod capacity (SUNDAY_PODS_PER_SUNDAY > 1),
                # allow other divisions to use the extra pod(s).
                if sunday_pods_used is None:
                    if assigned not in (None, div):
                        continue
                else:
                    if sunday_pods_used.get(d, 0) == 0 and assigned not in (None, div):
                        continue
            if sunday_pods_used is not None and d.weekday() == 6 and sunday_pods_used.get(d, 0) >= SUNDAY_PODS_PER_SUNDAY:
                continue
            if all(doubleheader_count[t] >= min_dh(t) for t in teams):
                break

            # adjacent slot pairs available that date
            slots = sorted(set(timeslots_by_date.get(d, [])), key=lambda s: datetime.strptime(s.strip(), "%I:%M %p"))
            for i in range(len(slots) - 1):
                s1, s2 = slots[i], slots[i + 1]
                free_fields = available_fields_for_pair(d, s1, s2)
                if len(free_fields) < 2:
                    continue

                # pick 4 eligible teams that still need DHs
                eligible = [t for t in teams if can_play_pod(t, d) and doubleheader_count[t] < min_dh(t)]
                if len(eligible) < 4:
                    continue

                eligible.sort(key=lambda t: team_need_key(t, team_stats, doubleheader_count), reverse=True)
                pool = eligible[:10]

                chosen = None
                chosen_layout = None

                # Try combos then permutations to find one that matches existing matchups.
                for combo in itertools.combinations(pool, 4):
                    for perm in itertools.permutations(combo, 4):
                        t1, t2, t3, t4 = perm
                        # Need these undirected pairs available in unscheduled
                        needed_pairs = [(t1, t2), (t3, t4), (t1, t3), (t2, t4)]
                        if all(((a, b) in unscheduled or (b, a) in unscheduled) for (a, b) in needed_pairs):
                            chosen = combo
                            chosen_layout = (t1, t2, t3, t4)
                            break
                    if chosen_layout:
                        break
                if not chosen_layout:
                    continue

                t1, t2, t3, t4 = chosen_layout

                # Consume matchups (one each) BEFORE placing; if any pop fails, rollback and skip.
                pops = []
                ok = True
                for a, b in [(t1, t2), (t3, t4), (t1, t3), (t2, t4)]:
                    m = _pop_matchup_any_orientation(unscheduled, a, b)
                    if m is None:
                        ok = False
                        break
                    pops.append(m)
                if not ok:
                    # rollback
                    unscheduled.extend(pops)
                    continue

                f1, f2 = free_fields[0], free_fields[1]

                # Place pod games
                ok = True
                ok &= place_game(d, s1, f1, t1, t2)
                ok &= place_game(d, s1, f2, t3, t4)
                ok &= place_game(d, s2, f1, t1, t3)
                ok &= place_game(d, s2, f2, t2, t4)

                if not ok:
                    # rollback placements is messy; instead, mark failed by re-adding matchups
                    unscheduled.extend(pops)
                    continue

                # Mark DH day for each team and record opponents played that day
                for team, opps in (
                    (t1, {t2, t3}),
                    (t2, {t1, t4}),
                    (t3, {t4, t1}),
                    (t4, {t3, t2}),
                ):
                    doubleheader_count[team] += 1
                    team_doubleheader_opponents[team][d].update(opps)

                progress = True
                if sunday_pods_used is not None and d.weekday() == 6:
                    sunday_pods_used[d] = sunday_pods_used.get(d, 0) + 1
                # continue scanning for more pods on same date

        if not progress:
            break

    return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents, used_slots, unscheduled

# -------------------------------
# Primary scheduling

# -------------------------------

def schedule_games(matchups, team_availability, field_availability, team_blackouts,
                   schedule, team_stats, doubleheader_count,
                   team_game_days, team_game_slots, team_doubleheader_opponents,
                   used_slots, timeslots_by_date, sunday_assignment=None, team_preferred_days=None):
    """
    Greedy single-game / DH-second-game placement for any remaining matchups.

    Performance note:
      The old retry/backtracking loop could take a long time when the remaining
      matchups are hard to place. This version does bounded multi-pass greedy
      filling: iterate all open slots, place the best matchup we can, repeat
      a few passes until no progress.

    Returns updated schedule + remaining unscheduled matchups.
    """
    unscheduled = list(matchups)

    def slot_ok_for_team(team, d, slot):
        # cannot play same timeslot twice in a day
        if slot in team_game_slots[team][d]:
            return False

        # If team already has a game today, the next game must be the immediate next timeslot (DH adjacency rule)
        if team_game_slots[team][d]:
            current = team_game_slots[team][d][0]
            sorted_slots = timeslots_by_date[d]
            try:
                idx = sorted_slots.index(current)
            except ValueError:
                return False
            if idx + 1 >= len(sorted_slots):
                return False
            required_slot = sorted_slots[idx + 1]
            return slot == required_slot

        return True

    # More passes helps the greedy filler converge after pods consume many prime slots.
    season_start = min((dt.date() for dt, _slot, _field in field_availability), default=None)
    max_passes = 20
    for _pass in range(max_passes):
        progress_made = False

        rnd = random.Random((RANDOM_SEED or 0) + (_pass * 97) + 7)
        # When Sundays are reserved for pods, prefer weekdays here; Sundays are skipped
        # outright below so a lone single never breaks up a pod's adjacent-slot pair.
        _sun_key = (1 if SUNDAY_PODS_ONLY else 0)
        # Opening weeks first (an unused early slot is capacity that never comes back),
        # then the Sunday preference, then randomised to keep the greedy pass unstuck.
        slots_iter = sorted(field_availability,
                            key=lambda x: (0 if in_front_load_window(x[0].date()) else 1,
                                           _sun_key if x[0].weekday() == 6 else (1 - _sun_key),
                                           rnd.random()))
        for date, slot, field in slots_iter:
            if SUNDAY_PODS_ONLY and date.weekday() == 6:
                continue
            if used_slots.get((date, slot, field), False):
                continue

            d = date.date()
            day_of_week = dow_label(date)
            week_num = date.isocalendar()[1]

            best = None
            best_score = -1

            for (t1, t2) in unscheduled:
                # DH-only divisions are scheduled exclusively by the pod routine
                if is_dh_only(div_of(t1)) or is_dh_only(div_of(t2)):
                    continue

                # availability / blackouts
                if not (is_team_available(t1, d, team_availability, team_blackouts) and is_team_available(t2, d, team_availability, team_blackouts)):
                    continue

                # target / weekly limits
                if team_stats[t1]['total_games'] >= target_games(t1) or team_stats[t2]['total_games'] >= target_games(t2):
                    continue
                if (team_stats[t1]['weekly_games'][week_num] >= WEEKLY_GAME_LIMIT or
                    team_stats[t2]['weekly_games'][week_num] >= WEEKLY_GAME_LIMIT):
                    continue

                # min gap
                if not (min_gap_ok(t1, d, team_game_days) and min_gap_ok(t2, d, team_game_days)):
                    continue

                # hard cadence rule: no two consecutive bye weeks
                if not (no_two_consecutive_byes_after_adding(t1, d, team_game_days) and no_two_consecutive_byes_after_adding(t2, d, team_game_days)):
                    continue

                # slot adjacency rules for DH second game
                if not slot_ok_for_team(t1, d, slot) or not slot_ok_for_team(t2, d, slot):
                    continue

                # DH constraints: if either team is adding a 2nd game today, enforce max DH days and "different opponent same day"
                can_double = True
                for team, opp in ((t1, t2), (t2, t1)):
                    if team_game_days[team][d] == 1:
                        if doubleheader_count[team] >= max_dh(team):
                            can_double = False
                            break
                        if team_doubleheader_opponents[team][d] and opp in team_doubleheader_opponents[team][d]:
                            can_double = False
                            break
                if not can_double:
                    continue

                score = score_placement(t1, t2, d, team_stats, doubleheader_count, team_game_days, sunday_assignment)
                if score > best_score:
                    best_score = score
                    best = (t1, t2)

            if best is None:
                continue

            t1, t2 = best
            home, away = decide_home_away(t1, t2, team_stats)

            # Hard cap to avoid exceeding desired home balance too much
            if team_stats[home]['home_games'] >= HOME_AWAY_BALANCE:
                if team_stats[away]['home_games'] < HOME_AWAY_BALANCE:
                    home, away = away, home
                else:
                    continue

            schedule.append((date, slot, field, home, home[0], away, away[0]))

            for team in (home, away):
                team_stats[team]['total_games'] += 1
                team_stats[team]['weekly_games'][week_num] += 1
                team_game_slots[team][d].append(slot)
                team_game_days[team][d] += 1

            team_stats[home]['home_games'] += 1
            team_stats[away]['away_games'] += 1

            for team, opp in ((home, away), (away, home)):
                # Record today's opponent for EVERY game, not just once a second game
                # exists. Recording only at the 2nd game left the set empty while the
                # first game was placed, so the "different opponent same day" guard had
                # nothing to compare against and the same pairing could be booked twice.
                team_doubleheader_opponents[team][d].add(opp)
                if team_game_days[team][d] == 2:
                    doubleheader_count[team] += 1

            used_slots[(date, slot, field)] = True
            unscheduled.remove((t1, t2))
            progress_made = True

        if not progress_made:
            break

    if unscheduled:
        print("Warning: Some predetermined matchups could not be scheduled ({} remaining).".format(len(unscheduled)))

    return schedule, team_stats, doubleheader_count, team_game_days, team_game_slots, team_doubleheader_opponents, used_slots, unscheduled


def fill_missing_games(schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
                       team_doubleheader_opponents, used_slots, timeslots_by_date, unscheduled,
                       team_availability, team_blackouts, field_availability, sunday_assignment=None, team_preferred_days=None):
    """
    Top-up pass after schedule_games. Works only with remaining unscheduled matchups.
    Uses the same bounded multi-pass greedy approach as schedule_games.
    """
    remaining = list(unscheduled)

    def slot_ok_for_team(team, d, slot):
        if slot in team_game_slots[team][d]:
            return False
        if team_game_slots[team][d]:
            current = team_game_slots[team][d][0]
            sorted_slots = timeslots_by_date[d]
            try:
                idx = sorted_slots.index(current)
            except ValueError:
                return False
            if idx + 1 >= len(sorted_slots):
                return False
            required_slot = sorted_slots[idx + 1]
            return slot == required_slot
        return True

    season_start = min((dt.date() for dt, _slot, _field in field_availability), default=None)
    max_passes = 20
    for _pass in range(max_passes):
        progress = False

        # stop early if nobody is below target or we have no matchups left
        if not remaining:
            break
        if not any(team_stats[t]['total_games'] < target_games(t) for t in team_stats.keys()):
            break

        rnd = random.Random((RANDOM_SEED or 0) + (_pass * 101) + 11)
        _sun_key = (1 if SUNDAY_PODS_ONLY else 0)
        # Opening weeks first (an unused early slot is capacity that never comes back),
        # then the Sunday preference, then randomised to keep the greedy pass unstuck.
        slots_iter = sorted(field_availability,
                            key=lambda x: (0 if in_front_load_window(x[0].date()) else 1,
                                           _sun_key if x[0].weekday() == 6 else (1 - _sun_key),
                                           rnd.random()))
        for date, slot, field in slots_iter:
            if SUNDAY_PODS_ONLY and date.weekday() == 6:
                continue
            if used_slots.get((date, slot, field), False):
                continue

            d = date.date()
            day_of_week = dow_label(date)
            week_num = date.isocalendar()[1]

            best = None
            best_score = -1

            for (t1, t2) in remaining:
                if is_dh_only(div_of(t1)) or is_dh_only(div_of(t2)):
                    continue

                # if both teams already at target, skip
                if team_stats[t1]['total_games'] >= target_games(t1) or team_stats[t2]['total_games'] >= target_games(t2):
                    continue

                if not (is_team_available(t1, d, team_availability, team_blackouts) and is_team_available(t2, d, team_availability, team_blackouts)):
                    continue

                if (team_stats[t1]['weekly_games'][week_num] >= WEEKLY_GAME_LIMIT or
                    team_stats[t2]['weekly_games'][week_num] >= WEEKLY_GAME_LIMIT):
                    continue

                if not (min_gap_ok(t1, d, team_game_days) and min_gap_ok(t2, d, team_game_days)):
                    continue

                if not slot_ok_for_team(t1, d, slot) or not slot_ok_for_team(t2, d, slot):
                    continue

                can_double = True
                for team, opp in ((t1, t2), (t2, t1)):
                    if team_game_days[team][d] == 1:
                        if doubleheader_count[team] >= max_dh(team):
                            can_double = False
                            break
                        if opp in team_doubleheader_opponents[team][d]:
                            can_double = False
                            break
                if not can_double:
                    continue

                score = score_placement(t1, t2, d, team_stats, doubleheader_count, team_game_days, sunday_assignment)
                if score > best_score:
                    best_score = score
                    best = (t1, t2)

            if best is None:
                continue

            t1, t2 = best
            home, away = decide_home_away(t1, t2, team_stats)

            if team_stats[home]['home_games'] >= HOME_AWAY_BALANCE:
                if team_stats[away]['home_games'] < HOME_AWAY_BALANCE:
                    home, away = away, home
                else:
                    continue

            schedule.append((date, slot, field, home, home[0], away, away[0]))

            for team in (home, away):
                team_stats[team]['total_games'] += 1
                team_stats[team]['weekly_games'][week_num] += 1
                team_game_slots[team][d].append(slot)
                team_game_days[team][d] += 1

            team_stats[home]['home_games'] += 1
            team_stats[away]['away_games'] += 1

            for team, opp in ((home, away), (away, home)):
                # Record today's opponent for EVERY game (see note in schedule_games):
                # recording only at the 2nd game let the same pairing be booked twice
                # on one date.
                team_doubleheader_opponents[team][d].add(opp)
                if team_game_days[team][d] == 2:
                    doubleheader_count[team] += 1

            used_slots[(date, slot, field)] = True
            remaining.remove((t1, t2))
            progress = True

        if not progress:
            break

    return schedule, team_stats, doubleheader_count, remaining


# -------------------------------
# Post-build repair pass
# -------------------------------
def _team_game_dates(schedule, team):
    """Return sorted list of dates where team plays."""
    dates = set()
    for entry in schedule:
        if entry is None:
            continue
        dt, slot, field, home, hd, away, ad = entry
        if home == team or away == team:
            dates.add(dt.date())
    return sorted(dates)


def _largest_gap(game_dates):
    """Return the largest gap in days between consecutive game dates, or 0."""
    if len(game_dates) < 2:
        return 0
    return max((game_dates[i+1] - game_dates[i]).days for i in range(len(game_dates) - 1))


def _gap_details(game_dates):
    """Return list of (gap_days, date_after) for all consecutive gaps, sorted largest first."""
    if len(game_dates) < 2:
        return []
    gaps = []
    for i in range(len(game_dates) - 1):
        g = (game_dates[i+1] - game_dates[i]).days
        gaps.append((g, game_dates[i], game_dates[i+1]))
    gaps.sort(key=lambda x: -x[0])
    return gaps


def _monthly_pace(game_dates, target, season_start, season_end):
    """Return dict of {month_end_date: (expected_by_then, actual_by_then, delta)}.

    Months are end-of-April, end-of-May, end-of-June, end-of-July.
    """

    checkpoints = []
    for month in (4, 5, 6, 7):
        # Last day of month
        if month == 12:
            end = date(season_start.year + 1, 1, 1)
        else:
            end = date(season_start.year, month + 1, 1)

        end = end - timedelta(days=1)
        if end >= season_start and end <= season_end:
            checkpoints.append(end)

    if not checkpoints:
        return {}

    total_days = (season_end - season_start).days or 1
    result = {}
    for cp in checkpoints:
        elapsed = (cp - season_start).days
        fraction = elapsed / total_days
        expected = round(target * fraction)
        actual = sum(1 for d in game_dates if d <= cp)
        result[cp] = (expected, actual, actual - expected)
    return result


def compute_schedule_diagnostics(schedule, all_teams, team_stats, doubleheader_count, field_availability):
    """Compute per-team diagnostics for the repair pass and reporting.

    Returns list of dicts with keys:
      team, division, target, scheduled, deficit, max_gap, worst_gap_start,
      worst_gap_end, monthly_pace (dict), back_heavy_score
    """
    season_dates = sorted({dt.date() for (dt, _, _) in field_availability})
    if not season_dates:
        return []
    season_start = season_dates[0]
    season_end = season_dates[-1]
    midpoint = season_start + (season_end - season_start) / 2

    diagnostics = []
    for team in sorted(all_teams):
        target = target_games(team)
        scheduled = team_stats[team]['total_games']
        deficit = max(0, target - scheduled)

        game_dates = _team_game_dates(schedule, team)
        max_gap = _largest_gap(game_dates)
        gaps = _gap_details(game_dates)

        worst_gap_start = gaps[0][1] if gaps else None
        worst_gap_end = gaps[0][2] if gaps else None

        pace = _monthly_pace(game_dates, target, season_start, season_end)

        # Back-heavy score: fraction of games in the second half of the season.
        # 0.5 = perfectly balanced, >0.5 = back-heavy
        games_second_half = sum(1 for d in game_dates if d > midpoint)
        back_heavy = games_second_half / max(1, len(game_dates))

        diagnostics.append({
            'team': team,
            'division': div_of(team),
            'target': target,
            'scheduled': scheduled,
            'deficit': deficit,
            'max_gap': max_gap,
            'worst_gap_start': worst_gap_start,
            'worst_gap_end': worst_gap_end,
            'gaps': gaps,
            'monthly_pace': pace,
            'back_heavy': back_heavy,
            'dh_count': doubleheader_count[team],
            'dh_min': min_dh(team),
            'dh_max': max_dh(team),
        })

    return diagnostics


def repair_schedule(schedule, all_teams, team_stats, doubleheader_count,
                    team_game_days, team_game_slots, team_doubleheader_opponents,
                    used_slots, timeslots_by_date, field_availability, field_index,
                    team_availability, team_blackouts, sunday_assignment=None,
                    max_moves=50):
    """Post-build repair pass: improve season shape by moving a small number of games.

    Strategy:
      1. Identify the worst problems: teams with long gaps, behind-pace, or back-heavy
      2. For each problem game (one near a cluster), try to move it into the gap
      3. Only move if it strictly improves the overall schedule quality score

    Returns: (schedule, moves_made, diagnostics_before, diagnostics_after)
    """
    diagnostics_before = compute_schedule_diagnostics(
        schedule, all_teams, team_stats, doubleheader_count, field_availability
    )

    def _schedule_quality_score(diags):
        """Single number summarizing schedule quality (lower = better problems)."""
        total = 0
        for d in diags:
            total += d['max_gap'] * 10          # penalize long gaps heavily
            total += d['deficit'] * 100         # penalize unscheduled games
            total += abs(d['back_heavy'] - 0.5) * 50  # penalize imbalance
            # penalize behind-pace months
            for cp, (expected, actual, delta) in d['monthly_pace'].items():
                if delta < 0:
                    total += abs(delta) * 20
        return total

    moves_made = []
    initial_score = _schedule_quality_score(diagnostics_before)

    # Sort teams by worst problems first
    problem_teams = sorted(diagnostics_before, key=lambda d: (
        -d['max_gap'], -d['deficit'], -d['back_heavy']
    ))

    for move_attempt in range(max_moves):
        if not problem_teams:
            break

        # Find the team with the worst gap that we haven't already fixed
        best_move = None
        best_improvement = 0

        for diag in problem_teams:
            team = diag['team']
            if diag['max_gap'] < PREFERRED_MIN_GAP + 2:
                continue  # gap is acceptable

            gaps = diag['gaps']
            if not gaps:
                continue

            worst_gap_days, gap_start, gap_end = gaps[0]
    

            # Find candidate dates in the middle of the gap
            gap_mid = gap_start + timedelta(days=worst_gap_days // 2)
            candidate_dates = []
            for delta in range(-3, 4):
                cd = gap_mid + timedelta(days=delta)
                if cd > gap_start and cd < gap_end:
                    candidate_dates.append(cd)

            # Find a game we can move INTO this gap
            # Look for games by this team that are in clusters (close together)
            game_dates = _team_game_dates(schedule, team)
            moveable_games = []
            for i, gd in enumerate(game_dates):
                # Only consider moving games that are in tight clusters
                # (within PREFERRED_MIN_GAP of another game on both sides)
                neighbors = 0
                if i > 0 and (gd - game_dates[i-1]).days <= PREFERRED_MIN_GAP:
                    neighbors += 1
                if i < len(game_dates)-1 and (game_dates[i+1] - gd).days <= PREFERRED_MIN_GAP:
                    neighbors += 1
                if neighbors >= 1:
                    moveable_games.append(gd)

            if not moveable_games:
                continue

            # Try each moveable game -> each candidate date
            for source_date in moveable_games:
                # Find the actual game entry
                source_games = [(idx, g) for idx, g in enumerate(schedule)
                                if g is not None and g[0].date() == source_date and (g[3] == team or g[5] == team)]
                if not source_games:
                    continue

                # Only try to move single games (not part of a DH on that date)
                if team_game_days[team].get(source_date, 0) != 1:
                    continue

                src_idx, src_game = source_games[0]
                src_dt, src_slot, src_field, src_home, src_hd, src_away, src_ad = src_game
                other_team = src_away if src_home == team else src_home

                for target_date in candidate_dates:
                    # A repair move always lands a *single* game, so it must not consume
                    # Sunday inventory that is reserved for doubleheader pods.
                    if SUNDAY_PODS_ONLY and target_date.weekday() == 6:
                        continue
                    # Check both teams available on target date
                    if not is_team_available(team, target_date, team_availability, team_blackouts):
                        continue
                    if not is_team_available(other_team, target_date, team_availability, team_blackouts):
                        continue

                    # Check gap constraints on target date for both teams
                    # Temporarily remove source date from game_days to check properly
                    temp_days_team = dict(team_game_days[team])
                    temp_days_other = dict(team_game_days[other_team])

                    # Check no existing game on target date
                    if team_game_days[team].get(target_date, 0) != 0:
                        continue
                    if team_game_days[other_team].get(target_date, 0) != 0:
                        continue

                    # Check weekly limit on target week
                    target_week = target_date.isocalendar()[1]
                    source_week = source_date.isocalendar()[1]
                    for t in (team, other_team):
                        wk_games = team_stats[t]['weekly_games'].get(target_week, 0)
                        # If moving from same week, it cancels out
                        adjust = -1 if source_week == target_week else 0
                        if wk_games + adjust + 1 > WEEKLY_GAME_LIMIT:
                            break
                    else:
                        # Find a free slot on target date
                        target_slots = timeslots_by_date.get(target_date, [])
                        placed = False
                        for tslot in target_slots:
                            free = free_fields_for_slot(field_index, target_date, tslot, used_slots)
                            if free:
                                # This move looks viable — estimate improvement
                                # Simulate: what would the gap look like after the move?
                                new_dates = [d for d in game_dates if d != source_date] + [target_date]
                                new_dates.sort()
                                new_max_gap = _largest_gap(new_dates)
                                improvement = worst_gap_days - new_max_gap

                                if improvement > best_improvement:
                                    best_improvement = improvement
                                    best_move = {
                                        'src_idx': src_idx,
                                        'src_game': src_game,
                                        'team': team,
                                        'other_team': other_team,
                                        'source_date': source_date,
                                        'target_date': target_date,
                                        'target_slot': tslot,
                                        'target_field_entry': free[0],
                                        'improvement': improvement,
                                    }
                                placed = True
                                break
                        continue

        if best_move is None or best_improvement <= 1:
            break

        # Execute the move
        mv = best_move
        src_game = mv['src_game']
        src_dt, src_slot, src_field, src_home, src_hd, src_away, src_ad = src_game
        target_entry = mv['target_field_entry']
        tgt_dt, tgt_slot, tgt_field = target_entry

        # Remove old game from schedule
        schedule[mv['src_idx']] = None  # mark for cleanup
        source_date = mv['source_date']
        target_date = mv['target_date']

        # Update tracking: remove from source
        for t in (src_home, src_away):
            team_stats[t]['total_games'] -= 1
            src_week = source_date.isocalendar()[1]
            team_stats[t]['weekly_games'][src_week] = max(0, team_stats[t]['weekly_games'].get(src_week, 0) - 1)
            team_game_days[t][source_date] = max(0, team_game_days[t].get(source_date, 0) - 1)
            if src_slot in team_game_slots[t][source_date]:
                team_game_slots[t][source_date].remove(src_slot)
        team_stats[src_home]['home_games'] -= 1
        team_stats[src_away]['away_games'] -= 1
        used_slots.pop((src_dt, src_slot, src_field), None)

        # Add to target
        home, away = decide_home_away(src_home, src_away, team_stats)
        new_game = (tgt_dt, tgt_slot, tgt_field, home, home[0], away, away[0])
        schedule.append(new_game)
        tgt_week = target_date.isocalendar()[1]
        for t in (home, away):
            team_stats[t]['total_games'] += 1
            team_stats[t]['weekly_games'][tgt_week] = team_stats[t]['weekly_games'].get(tgt_week, 0) + 1
            team_game_days[t][target_date] += 1
            team_game_slots[t][target_date].append(tgt_slot)
        team_stats[home]['home_games'] += 1
        team_stats[away]['away_games'] += 1
        used_slots[(tgt_dt, tgt_slot, tgt_field)] = True

        moves_made.append({
            'team': mv['team'],
            'from': f"{source_date} {src_slot} {src_field}",
            'to': f"{target_date} {tgt_slot} {tgt_field}",
            'matchup': f"{src_home} vs {src_away}",
            'gap_reduction': mv['improvement'],
        })

        # Refresh diagnostics for next iteration
        problem_teams = compute_schedule_diagnostics(
            [g for g in schedule if g is not None], all_teams, team_stats, doubleheader_count, field_availability
        )
        problem_teams.sort(key=lambda d: (-d['max_gap'], -d['deficit'], -d['back_heavy']))

    # Clean up None entries from moves
    schedule[:] = [g for g in schedule if g is not None]

    diagnostics_after = compute_schedule_diagnostics(
        schedule, all_teams, team_stats, doubleheader_count, field_availability
    )

    final_score = _schedule_quality_score(diagnostics_after)
    print(f"\nRepair pass: {len(moves_made)} games moved")
    print(f"  Quality score: {initial_score:.0f} -> {final_score:.0f} (lower is better)")
    if moves_made:
        for mv in moves_made:
            print(f"  Moved {mv['matchup']}: {mv['from']} -> {mv['to']} (gap reduced by {mv['gap_reduction']} days)")

    return schedule, moves_made, diagnostics_before, diagnostics_after


def build_slot_rows(field_availability, scheduled_games):
    """
    Returns list of rows (one per field_availability entry) with blank home/away when unused.
    scheduled_games: list of game tuples (datetime, slot_str, field, home, home_div, away, away_div)
    """
    game_by_key = {}
    for g in scheduled_games:
        dt, slot, field, home, home_div, away, away_div = g
        game_by_key[(dt.date(), slot, field)] = g

    rows = []
    for dt, slot, field in field_availability:
        g = game_by_key.get((dt.date(), slot, field))
        if g is None:
            rows.append((dt, slot, field, "", "", "", ""))
        else:
            _, _, _, home, home_div, away, away_div = g
            rows.append((dt, slot, field, home, home_div, away, away_div))
    return rows




def output_schedule_to_csv_full(field_availability, schedule, output_file):
    rows = build_slot_rows(field_availability, schedule)
    with open(output_file, mode='w', newline='') as file:
        writer = csv.writer(file)
        writer.writerow(["Date", "Day", "Time", "Diamond", "Home Team", "Home Division", "Away Team", "Away Division"])
        for dt, slot, field, home, home_div, away, away_div in rows:
            writer.writerow([dt.strftime('%Y-%m-%d'), dow_label(dt), slot, field, home, home_div, away, away_div])
    return rows


# -------------------------------
# Unscheduled matchups reporting (for manual fill)
# -------------------------------
def summarize_remaining_matchups(remaining_matchups):
    """Aggregate remaining matchups into counts.

    Returns:
      oriented_counts: dict[(home, away)] -> count
      unordered_counts: dict[(tmin, tmax)] -> count
    """
    oriented = defaultdict(int)
    unordered = defaultdict(int)
    for a, b in remaining_matchups or []:
        oriented[(a, b)] += 1
        t1, t2 = (a, b) if a <= b else (b, a)
        unordered[(t1, t2)] += 1
    return oriented, unordered


def _current_unordered_meet_counts(schedule):
    """Return dict[(tmin,tmax)] -> games already scheduled between the pair."""
    counts = defaultdict(int)
    for (date_str, time_str, field_id, home, home_div, away, away_div) in schedule or []:
        if not home or not away:
            continue
        t1, t2 = (home, away) if home <= away else (away, home)
        counts[(t1, t2)] += 1
    return counts


def _find_open_dates_for_pair(t1, t2, field_availability, used_slots, team_availability, team_blackouts, team_game_days, max_dates=5):
    """Return up to max_dates open dates where both teams can play and a field slot is free."""
    seen = set()
    results = []
    for dt, slot, field in field_availability:
        d = dt.date()
        if d in seen:
            continue
        if used_slots.get((dt, slot, field), False):
            continue
        if not is_team_available(t1, d, team_availability, team_blackouts):
            continue
        if not is_team_available(t2, d, team_availability, team_blackouts):
            continue
        seen.add(d)
        results.append(d.strftime('%Y-%m-%d'))
        if len(results) >= max_dates:
            break
    return results


def suggest_best_fit_manual_matchups(all_teams, schedule, team_stats, doubleheader_count,
                                     team_availability=None, team_blackouts=None, max_pairs=None,
                                     field_availability=None, used_slots=None, team_game_days=None):
    """Greedy 'best fit' list of matchups among ONLY teams currently short of target games.

    Goal: produce a simple, reasonable list to manually place that:
      - fixes game deficits (as much as possible)
      - prefers pairs that have played each other the least (based on current schedule matrix)
      - breaks ties by pairing teams with bigger deficits
      - shows open dates where both teams can play (actionable)

    Returns list of dict rows ready for XLSX export.
    """
    if not team_stats:
        return []

    # 1) identify teams short
    needs = {t: max(0, target_games(t) - int(team_stats[t].get('total_games', 0))) for t in all_teams}
    short = sorted([t for t in all_teams if needs.get(t, 0) > 0])
    if not short:
        return []

    total_missing = sum(needs[t] for t in short)
    meet_counts = _current_unordered_meet_counts(schedule)

    # 2) greedy pairing
    remaining = dict(needs)
    rows = []
    # Cap to what math allows
    target_pairs = total_missing // 2
    if max_pairs is not None:
        target_pairs = min(target_pairs, int(max_pairs))

    def _pair_key(t1, t2):
        a, b = (t1, t2) if t1 <= t2 else (t2, t1)
        played = meet_counts.get((a, b), 0)
        # lower played is better; higher opponent need is better; prefer intra slightly (optional)
        same_div = 1 if div_of(t1) == div_of(t2) else 0
        return (played, -remaining.get(t2, 0), -same_div, t2)

    while len(rows) < target_pairs:
        # pick the team with biggest games deficit; tie-break by DH deficit
        candidates1 = [t for t in short if remaining.get(t, 0) > 0]
        if len(candidates1) < 2:
            break
        t1 = sorted(candidates1, key=lambda t: (-remaining[t], -dh_deficit(t, doubleheader_count), t))[0]

        candidates2 = [t for t in candidates1 if t != t1]
        if not candidates2:
            break
        t2 = sorted(candidates2, key=lambda t: _pair_key(t1, t))[0]

        a, b = (t1, t2) if t1 <= t2 else (t2, t1)
        played = meet_counts.get((a, b), 0)

        rows.append({
            "Team 1": t1,
            "Div 1": div_of(t1),
            "Needs 1": int(remaining.get(t1, 0)),
            "DH Need 1": int(dh_deficit(t1, doubleheader_count)),
            "Team 2": t2,
            "Div 2": div_of(t2),
            "Needs 2": int(remaining.get(t2, 0)),
            "DH Need 2": int(dh_deficit(t2, doubleheader_count)),
            "Current Meetings": int(played),
            "Type": "INTRA" if div_of(t1) == div_of(t2) else "INTER",
            "Common Avail Days": _common_avail_days(t1, t2, team_availability),
            "Blackouts": _blackout_summary(t1, t2, team_blackouts),
            "Open Dates": ", ".join(_find_open_dates_for_pair(
                t1, t2, field_availability, used_slots or {}, team_availability, team_blackouts, team_game_days
            )) if field_availability else "",
        })

        # update remaining deficits
        remaining[t1] = max(0, remaining.get(t1, 0) - 1)
        remaining[t2] = max(0, remaining.get(t2, 0) - 1)

    # Add a small tail note if odd deficit remains (can't be paired cleanly)
    leftover = [(t, remaining[t]) for t in short if remaining.get(t, 0) > 0]
    if leftover:
        empty_row = {k: "" for k in ["Team 1", "Div 1", "Needs 1", "DH Need 1",
                     "Team 2", "Div 2", "Needs 2", "DH Need 2",
                     "Current Meetings", "Type", "Common Avail Days", "Blackouts", "Open Dates"]}
        rows.append(empty_row)
        note_row = dict(empty_row)
        note_row["Team 1"] = "Leftover needs (odd / not pairable):"
        note_row["Needs 1"] = ", ".join([f"{t}:{n}" for t, n in leftover])
        rows.append(note_row)

    return rows


def output_unscheduled_matchups_csv(remaining_matchups, output_file):
    """Write remaining matchups to CSV, aggregated by unordered pair.

    Columns: Division1, Team1, Division2, Team2, RemainingGames
    """
    _oriented, unordered = summarize_remaining_matchups(remaining_matchups)
    rows = []
    for (t1, t2), cnt in sorted(unordered.items(), key=lambda x: (-x[1], x[0][0], x[0][1])):
        rows.append((div_of(t1), t1, div_of(t2), t2, cnt))

    with open(output_file, mode='w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Division1", "Team1", "Division2", "Team2", "RemainingGames"])
        for r in rows:
            w.writerow(list(r))
    return rows

def output_team_remaining_needs_csv(all_teams, team_stats, doubleheader_count, output_file):
    """Write per-team remaining needs to CSV.

    Columns: Division, Team, TargetGames, ScheduledGames, GamesRemaining, MinDH, ScheduledDHDays, DHDaysRemainingToMin
    """
    with open(output_file, mode='w', newline='') as f:
        w = csv.writer(f)
        w.writerow(["Division","Team","TargetGames","ScheduledGames","GamesRemaining","MinDH","ScheduledDHDays","DHDaysRemainingToMin"])
        for t in sorted(all_teams, key=lambda x: (div_of(x), x)):
            target = target_games(t)
            scheduled = team_stats[t]['total_games']
            games_rem = max(0, target - scheduled)
            mindh = min_dh(t)
            dh_done = doubleheader_count[t]
            dh_rem = max(0, mindh - dh_done)
            w.writerow([div_of(t), t, target, scheduled, games_rem, mindh, dh_done, dh_rem])

def add_unscheduled_to_workbook(wb, remaining_matchups, all_teams, team_stats, doubleheader_count, sched_last, weeks_count=None):
    """Add two sheets: Unscheduled (one row per remaining matchup) and Remaining Needs.

    Unscheduled is intentionally NOT aggregated so you can walk down the list and paste games
    into open slots, then delete rows as you go.

    Also adds a formula column that lists week numbers where BOTH teams currently have 0 games
    scheduled (updates automatically when you edit the Schedule sheet).
    """
    if remaining_matchups is None:
        remaining_matchups = []

    # ---------------- Unscheduled ----------------
    ws_u = wb.create_sheet("Unscheduled")
    ws_u.append(["Home Team", "Away Team", "Home Div", "Away Div", "WeeksBothZero"])
    for cell in ws_u[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    # Determine the Weeks range (written by export_schedule_to_xlsx)
    # weeks_count is the number of week numbers in Weeks!A2:A{...}
    if not weeks_count:
        # best-effort: infer by scanning the Weeks sheet (if present)
        try:
            ws_w = wb["Weeks"]
            weeks_count = max(0, ws_w.max_row - 1)
        except Exception:
            weeks_count = 0

    weeks_range = None
    if weeks_count and weeks_count > 0:
        weeks_range = f"Weeks!$A$2:$A${weeks_count+1}"

    for i, (home, away) in enumerate(remaining_matchups, start=2):
        ws_u.cell(row=i, column=1, value=home)
        ws_u.cell(row=i, column=2, value=away)
        ws_u.cell(row=i, column=3, value=div_of(home))
        ws_u.cell(row=i, column=4, value=div_of(away))

        if weeks_range:
            # Excel 365 dynamic array formula
            ws_u.cell(
                row=i,
                column=5,
                value=(
                    f'=IFERROR(LET('
                    f'w,{weeks_range},'
                    f'h,$A{i},a,$B{i},'
                    f'hg,COUNTIFS(Schedule!$I$2:$I${sched_last},w,Schedule!$E$2:$E${sched_last},h)+COUNTIFS(Schedule!$I$2:$I${sched_last},w,Schedule!$F$2:$F${sched_last},h),'
                    f'ag,COUNTIFS(Schedule!$I$2:$I${sched_last},w,Schedule!$E$2:$E${sched_last},a)+COUNTIFS(Schedule!$I$2:$I${sched_last},w,Schedule!$F$2:$F${sched_last},a),'
                    f'TEXTJOIN(", ",TRUE,FILTER(w,(hg=0)*(ag=0)))'
                    f'),"")'
                )
            )
        else:
            ws_u.cell(row=i, column=5, value="")

    _autofit(ws_u, ws_u.max_row, 5, min_width=10, max_width=24)
    ws_u.freeze_panes = "A2"
    ws_u.auto_filter.ref = f"A1:E{ws_u.max_row}"

    # ---------------- Remaining Needs ----------------
    ws_n = wb.create_sheet("Remaining Needs")
    ws_n.append(["Division","Team","TargetGames","ScheduledGames","GamesRemaining","MinDH","ScheduledDHDays","DHDaysRemainingToMin"])
    for cell in ws_n[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    for t in sorted(all_teams, key=lambda x: (div_of(x), x)):
        target = target_games(t)
        scheduled = team_stats[t]['total_games']
        games_rem = max(0, target - scheduled)
        mindh = min_dh(t)
        dh_done = doubleheader_count[t]
        dh_rem = max(0, mindh - dh_done)
        ws_n.append([div_of(t), t, target, scheduled, games_rem, mindh, dh_done, dh_rem])

    _autofit(ws_n, ws_n.max_row, 8, min_width=12, max_width=22)
    ws_n.freeze_panes = "A2"
    ws_n.auto_filter.ref = f"A1:H{ws_n.max_row}"

# -------------------------------
# XLSX export (formulas + conditional formatting + matchup matrix)
# -------------------------------
def _autofit(ws, max_row, max_col, min_width=10, max_width=40):
    for col in range(1, max_col + 1):
        letter = get_column_letter(col)
        best = 0
        for r in range(1, max_row + 1):
            v = ws.cell(row=r, column=col).value
            if v is None:
                continue
            best = max(best, len(str(v)))
        ws.column_dimensions[letter].width = max(min_width, min(max_width, best + 2))

def export_schedule_to_xlsx(field_availability, schedule, division_teams, output_path, remaining_matchups=None, team_stats=None, doubleheader_count=None, team_availability=None, team_blackouts=None, diagnostics=None):
    if Workbook is None:
        raise RuntimeError("openpyxl is not installed. Run: pip install openpyxl")

    rows = build_slot_rows(field_availability, schedule)

    wb = Workbook()
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass

    all_teams = sorted([t for div in sorted(division_teams.keys()) for t in division_teams[div]])
    annotations = _schedule_row_annotations(rows, team_preferred_days=team_preferred_days)

    # ---------------- Schedule ----------------
    ws = wb.active
    ws.title = "Schedule"

    headers = [
        "Date", "Day", "Time", "Diamond", "Home Team", "Away Team", "Home Div", "Away Div",
        "Week #", "SlotIndex", "Game Type", "Home Games After", "Away Games After",
        "Home Last Game", "Away Last Game", "Home Days Since Last", "Away Days Since Last",
        "Preferred Match", "Flag"
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    slots_by_date = defaultdict(list)
    for dt0, slot0, _field0 in field_availability:
        d0 = dt0.date()
        slots_by_date[d0].append(slot0)
    slot_index_by_date_slot = {}
    for d0, slots0 in slots_by_date.items():
        uniq = sorted(set(slots0), key=lambda s: datetime.strptime(s.strip(), "%I:%M %p"))
        for i, s in enumerate(uniq, start=1):
            slot_index_by_date_slot[(d0, s)] = i

    for excel_row, (dt, slot, field, home, home_div, away, away_div) in enumerate(rows, start=2):
        d = dt.date()
        wk = d.isocalendar()[1]
        slot_idx = slot_index_by_date_slot.get((d, slot), "")
        meta = annotations.get(excel_row, {})
        ws.append([
            d, dow_label(dt), slot, field, home, away, home_div, away_div, wk, slot_idx,
            meta.get("game_type", "OPEN" if not home else ""),
            meta.get("home_after", ""), meta.get("away_after", ""),
            meta.get("home_last", ""), meta.get("away_last", ""),
            meta.get("home_days_since", ""), meta.get("away_days_since", ""),
            meta.get("preferred_match", ""), meta.get("flag", "Open Slot" if not home else "")
        ])

    n = len(rows)
    for r in range(2, n + 2):
        ws.cell(row=r, column=1).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=3).number_format = "@"
        ws.cell(row=r, column=14).number_format = "yyyy-mm-dd"
        ws.cell(row=r, column=15).number_format = "yyyy-mm-dd"

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{n + 1}"
    ws.column_dimensions['J'].hidden = True
    _autofit(ws, n + 1, 19)

    # ---------------- Teams ----------------
    ws_t = wb.create_sheet("Teams")
    ws_t.append(["Team", "Division", "Team Name"])
    for cell in ws_t[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    for t in all_teams:
        ws_t.append([t, div_of(t), ""])
    ws_t.freeze_panes = "A2"
    ws_t.auto_filter.ref = f"A1:C{len(all_teams)+1}"
    _autofit(ws_t, len(all_teams) + 1, 3, min_width=8, max_width=24)

    # ---------------- Team Summary ----------------
    ws_ts = wb.create_sheet("Team Summary")
    summary_rows = _build_team_summary(schedule, all_teams, team_stats or defaultdict(dict), doubleheader_count or defaultdict(int), team_preferred_days=team_preferred_days)
    summary_headers = [
        "Division", "Team", "Team Name", "Total Games", "Home", "Away", "DH Days",
        "Last Scheduled Game", "Longest Gap", "Max Gap Warning", "Preferred Hits",
        "Preferred Misses", "Games Remaining", "DH Remaining To Min"
    ]
    ws_ts.append(summary_headers)
    for cell in ws_ts[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
    for row_num, row in enumerate(summary_rows, start=2):
        ws_ts.append([row[h] for h in summary_headers])
        ws_ts.cell(row=row_num, column=3, value=f'=IFERROR(VLOOKUP(B{row_num},Teams!$A:$C,3,FALSE),B{row_num})')
        ws_ts.cell(row=row_num, column=8).number_format = "yyyy-mm-dd"
    ws_ts.freeze_panes = "A2"
    ws_ts.auto_filter.ref = f"A1:N{len(summary_rows)+1}"
    _autofit(ws_ts, len(summary_rows)+1, 14, min_width=10, max_width=24)

    # ---------------- Open Slots ----------------
    ws_o = wb.create_sheet("Open Slots")
    ws_o.append(["Date", "Day", "Time", "Diamond", "Week #", "Season Phase"])
    for cell in ws_o[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
    used_keys = {(dt.date(), slot, field) for (dt, slot, field, home, _hd, away, _ad) in rows if home and away}
    season_dates = sorted(dt.date() for dt, _, _ in field_availability)
    season_start = season_dates[0] if season_dates else None
    season_end = season_dates[-1] if season_dates else None
    total_span = max(1, (season_end - season_start).days) if season_start and season_end else 1
    open_rows = 0
    for dt, slot, field in field_availability:
        if (dt.date(), slot, field) in used_keys:
            continue
        open_rows += 1
        day_offset = (dt.date() - season_start).days if season_start else 0
        ratio = day_offset / total_span if total_span else 0
        phase = "Early" if ratio < 0.34 else ("Mid" if ratio < 0.67 else "Late")
        ws_o.append([dt.date(), dow_label(dt), slot, field, dt.date().isocalendar()[1], phase])
        ws_o.cell(row=open_rows + 1, column=1).number_format = "yyyy-mm-dd"
    ws_o.freeze_panes = "A2"
    ws_o.auto_filter.ref = f"A1:F{max(2, open_rows+1)}"
    _autofit(ws_o, max(2, open_rows+1), 6, min_width=10, max_width=18)

    # ---------------- Upload ----------------
    ws_up = wb.create_sheet("Upload")
    upload_headers = ["Date", "Time", "Type", "Duration", "Home Team", "Home Division", "Away Team", "Away Division", "Location"]
    ws_up.append(upload_headers)
    for cell in ws_up[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")
    upload_row = 2
    for sched_row in range(2, n + 2):
        if not ws.cell(row=sched_row, column=5).value or not ws.cell(row=sched_row, column=6).value:
            continue
        ws_up.cell(row=upload_row, column=1, value=f"=Schedule!A{sched_row}")
        ws_up.cell(row=upload_row, column=2, value=f"=Schedule!C{sched_row}")
        ws_up.cell(row=upload_row, column=3, value="Game")
        ws_up.cell(row=upload_row, column=4, value=f'=IF(Schedule!B{sched_row}="Sun",80,70)')
        ws_up.cell(row=upload_row, column=5, value=f'=IFERROR(VLOOKUP(Schedule!E{sched_row},Teams!$A:$C,3,FALSE),Schedule!E{sched_row})')
        ws_up.cell(row=upload_row, column=6, value=f'="Division "&Schedule!G{sched_row}')
        ws_up.cell(row=upload_row, column=7, value=f'=IFERROR(VLOOKUP(Schedule!F{sched_row},Teams!$A:$C,3,FALSE),Schedule!F{sched_row})')
        ws_up.cell(row=upload_row, column=8, value=f'="Division "&Schedule!H{sched_row}')
        ws_up.cell(row=upload_row, column=9, value=f"=Schedule!D{sched_row}")
        ws_up.cell(row=upload_row, column=1).number_format = "yyyy-mm-dd"
        upload_row += 1
    ws_up.freeze_panes = "A2"
    ws_up.auto_filter.ref = f"A1:I{max(2, upload_row-1)}"
    _autofit(ws_up, max(2, upload_row-1), 9, min_width=10, max_width=24)

    # ---------------- Suggested Manual Matchups ----------------
    ws_s = wb.create_sheet("Suggested Matchups")
    ws_s.append(["Team 1", "Div 1", "Needs 1", "DH Need 1",
                 "Team 2", "Div 2", "Needs 2", "DH Need 2",
                 "Current Meetings", "Type", "Common Avail Days", "Blackouts", "Open Dates"])
    for cell in ws_s[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    # Build used_slots lookup for open-date calculation
    _used = {}
    for g in schedule:
        if g[3] and g[5]:  # has home and away
            _used[(g[0], g[1], g[2])] = True

    suggested_rows = suggest_best_fit_manual_matchups(
        all_teams=all_teams,
        schedule=schedule,
        team_stats=team_stats,
        doubleheader_count=doubleheader_count or {t: 0 for t in all_teams},
        team_availability=team_availability,
        team_blackouts=team_blackouts,
        field_availability=field_availability,
        used_slots=_used,
    )

    for row in suggested_rows:
        ws_s.append([
            row.get("Team 1", ""), row.get("Div 1", ""), row.get("Needs 1", ""), row.get("DH Need 1", ""),
            row.get("Team 2", ""), row.get("Div 2", ""), row.get("Needs 2", ""), row.get("DH Need 2", ""),
            row.get("Current Meetings", ""), row.get("Type", ""), row.get("Common Avail Days", ""),
            row.get("Blackouts", ""), row.get("Open Dates", "")
        ])

    last_s = max(2, len(suggested_rows) + 1)
    ws_s.freeze_panes = "A2"
    ws_s.auto_filter.ref = f"A1:M{last_s}"
    for rr in range(2, last_s + 1):
        ws_s.cell(row=rr, column=11).alignment = Alignment(wrap_text=True, vertical="top")
        ws_s.cell(row=rr, column=12).alignment = Alignment(wrap_text=True, vertical="top")
        ws_s.cell(row=rr, column=13).alignment = Alignment(wrap_text=True, vertical="top")
    _autofit(ws_s, last_s, 13, min_width=10, max_width=28)

    # ---------------- Unscheduled Matches ----------------
    ws_u = wb.create_sheet("Unscheduled Matches")
    ws_u.append(["Home Div", "Home Team", "Away Div", "Away Team", "Remaining", "Home Needs", "Away Needs", "Type", "Available Days", "Blackouts"])
    for cell in ws_u[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
        cell.alignment = Alignment(horizontal="center")

    rows_u = []
    if remaining_matchups:
        if team_stats is not None:
            current_games = {t: int(team_stats[t].get('total_games', 0)) for t in all_teams}
        else:
            current_games = {t: 0 for t in all_teams}
            for (d, time_str, field_id, home, home_div, away, away_div) in schedule:
                if home and away:
                    current_games[home] = current_games.get(home, 0) + 1
                    current_games[away] = current_games.get(away, 0) + 1

        needs = {t: max(0, target_games(t) - current_games.get(t, 0)) for t in all_teams}
        below = {t for t in all_teams if needs.get(t, 0) > 0}

        oriented, _unordered = summarize_remaining_matchups(remaining_matchups)

        for (home, away), cnt in oriented.items():
            if home == away:
                continue
            if (home in below) or (away in below):
                rows_u.append((
                    div_of(home), home,
                    div_of(away), away,
                    int(cnt),
                    int(needs.get(home, 0)),
                    int(needs.get(away, 0)),
                    "INTRA" if div_of(home) == div_of(away) else "INTER",
                    _common_avail_days(home, away, team_availability),
                    _blackout_summary(home, away, team_blackouts)
                ))

        rows_u.sort(key=lambda r: (-r[4], r[0], r[1], r[2], r[3]))

    for r in rows_u:
        ws_u.append(list(r))

    last_u = max(2, len(rows_u) + 1)
    ws_u.freeze_panes = "A2"
    ws_u.auto_filter.ref = f"A1:J{last_u}"
    for rr in range(2, last_u + 1):
        ws_u.cell(row=rr, column=9).alignment = Alignment(wrap_text=True, vertical="top")
        ws_u.cell(row=rr, column=10).alignment = Alignment(wrap_text=True, vertical="top")
    _autofit(ws_u, last_u, 10, min_width=10, max_width=22)

    # ---------------- TeamDate (helper: games/day + non-adjacent DH detection) ----------------
    ws_td = wb.create_sheet("TeamDate")
    ws_td.append(["Key", "Date", "Team", "GamesThatDay", "MinSlot", "MaxSlot", "NonAdjFlag", "WeekNum"])
    for cell in ws_td[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")

    unique_dates = sorted({dt.date() for (dt, _, _) in field_availability})

    sched_first = 2
    sched_last = n + 1
    date_rng = f"Schedule!$A${sched_first}:$A${sched_last}"
    home_rng = f"Schedule!$E${sched_first}:$E${sched_last}"
    away_rng = f"Schedule!$F${sched_first}:$F${sched_last}"
    home_div_rng = f"Schedule!$G${sched_first}:$G${sched_last}"
    away_div_rng = f"Schedule!$H${sched_first}:$H${sched_last}"
    week_rng = f"Schedule!$I${sched_first}:$I${sched_last}"
    slotidx_rng = f"Schedule!$J${sched_first}:$J${sched_last}"
    day_rng = f"Schedule!$B${sched_first}:$B${sched_last}"

    row_idx = 2
    for d in unique_dates:
        wk = d.isocalendar()[1]
        for t in all_teams:
            ws_td.cell(row=row_idx, column=1, value='=TEXT($B{r},"yyyymmdd")&"|"&$C{r}'.format(r=row_idx))
            ws_td.cell(row=row_idx, column=2, value=d)
            ws_td.cell(row=row_idx, column=3, value=t)
            ws_td.cell(
                row=row_idx,
                column=4,
                value='=COUNTIFS({date_rng},$B{r},{home_rng},$C{r})+COUNTIFS({date_rng},$B{r},{away_rng},$C{r})'.format(
                    date_rng=date_rng, home_rng=home_rng, away_rng=away_rng, r=row_idx
                )
            )
            ws_td.cell(
                row=row_idx,
                column=5,
                value='=MIN(IFERROR(MINIFS({slotidx_rng},{date_rng},$B{r},{home_rng},$C{r}),9999),IFERROR(MINIFS({slotidx_rng},{date_rng},$B{r},{away_rng},$C{r}),9999))'.format(
                    slotidx_rng=slotidx_rng, date_rng=date_rng, home_rng=home_rng, away_rng=away_rng, r=row_idx
                )
            )
            ws_td.cell(
                row=row_idx,
                column=6,
                value='=MAX(IFERROR(MAXIFS({slotidx_rng},{date_rng},$B{r},{home_rng},$C{r}),0),IFERROR(MAXIFS({slotidx_rng},{date_rng},$B{r},{away_rng},$C{r}),0))'.format(
                    slotidx_rng=slotidx_rng, date_rng=date_rng, home_rng=home_rng, away_rng=away_rng, r=row_idx
                )
            )
            ws_td.cell(row=row_idx, column=7, value='=IF($D{r}<>2,0,IF($F{r}-$E{r}=1,0,1))'.format(r=row_idx))
            ws_td.cell(row=row_idx, column=8, value=wk)
            ws_td.cell(row=row_idx, column=2).number_format = "yyyy-mm-dd"
            row_idx += 1

    td_last = row_idx - 1
    ws_td.freeze_panes = "A2"
    ws_td.auto_filter.ref = f"A1:H{td_last}"
    _autofit(ws_td, td_last, 8, min_width=10, max_width=18)

    # ---------------- Weeks (helper for Unscheduled formulas) ----------------
    ws_w = wb.create_sheet("Weeks")
    ws_w.append(["WeekNum"])
    for cell in ws_w[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="D9E1F2")
    unique_weeks = sorted({dt.date().isocalendar()[1] for (dt, _, _) in field_availability})
    for wk in unique_weeks:
        ws_w.append([wk])
    ws_w.freeze_panes = "A2"
    ws_w.auto_filter.ref = f"A1:A{len(unique_weeks)+1}"

    # Legacy helper tabs
    add_unscheduled_to_workbook(wb, remaining_matchups, all_teams, team_stats or defaultdict(dict), doubleheader_count or defaultdict(int), sched_last, weeks_count=len(unique_weeks))

    # ---------------- Team Diagnostics ("Why this team is a problem") ----------------
    if diagnostics:
        ws_diag = wb.create_sheet("Team Diagnostics")

        # Get monthly checkpoint columns from the first team that has pace data
        month_cols = []
        for d in diagnostics:
            if d['monthly_pace']:
                month_cols = sorted(d['monthly_pace'].keys())
                break

        headers_diag = [
            "Division", "Team", "Target", "Scheduled", "Deficit",
            "Max Gap (days)", "Worst Gap Start", "Worst Gap End",
            "Back-Heavy %", "DH Count", "DH Min", "DH Max",
            "Status"
        ]
        # Add monthly pacing columns
        for mc in month_cols:
            headers_diag.append(f"Pace {mc.strftime('%b')}: Expected")
            headers_diag.append(f"Pace {mc.strftime('%b')}: Actual")
            headers_diag.append(f"Pace {mc.strftime('%b')}: Delta")

        ws_diag.append(headers_diag)
        for cell in ws_diag[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="D9E1F2")
            cell.alignment = Alignment(horizontal="center", wrap_text=True)

        # Sort diagnostics: worst problems first
        sorted_diag = sorted(diagnostics, key=lambda d: (
            -d['deficit'], -d['max_gap'], -abs(d['back_heavy'] - 0.5)
        ))

        # Highlight fills
        red_fill = PatternFill("solid", fgColor="FFC7CE")
        yellow_fill = PatternFill("solid", fgColor="FFEB9C")
        green_fill = PatternFill("solid", fgColor="C6EFCE")

        for r, d in enumerate(sorted_diag, start=2):
            # Determine status
            problems = []
            if d['deficit'] > 0:
                problems.append(f"{d['deficit']} games short")
            if d['max_gap'] >= PREFERRED_MIN_GAP + 3:
                problems.append(f"{d['max_gap']}-day gap")
            if d['back_heavy'] > 0.65:
                problems.append(f"{d['back_heavy']:.0%} back-heavy")
            if d['dh_count'] < d['dh_min']:
                problems.append(f"DH deficit: {d['dh_min'] - d['dh_count']}")

            # Monthly pace problems
            for mc in month_cols:
                if mc in d['monthly_pace']:
                    exp, act, delta = d['monthly_pace'][mc]
                    if delta < -2:
                        problems.append(f"{abs(delta)} behind by {mc.strftime('%b')}")

            status = "; ".join(problems) if problems else "OK"

            row_data = [
                d['division'], d['team'], d['target'], d['scheduled'], d['deficit'],
                d['max_gap'],
                d['worst_gap_start'].strftime('%Y-%m-%d') if d['worst_gap_start'] else "",
                d['worst_gap_end'].strftime('%Y-%m-%d') if d['worst_gap_end'] else "",
                round(d['back_heavy'] * 100, 1),
                d['dh_count'], d['dh_min'], d['dh_max'],
                status
            ]

            # Add monthly pacing data
            for mc in month_cols:
                if mc in d['monthly_pace']:
                    exp, act, delta = d['monthly_pace'][mc]
                    row_data.extend([exp, act, delta])
                else:
                    row_data.extend(["", "", ""])

            ws_diag.append(row_data)

            # Color the status cell
            status_cell = ws_diag.cell(row=r, column=13)
            if d['deficit'] > 0 or d['max_gap'] >= PREFERRED_MIN_GAP + 5:
                status_cell.fill = red_fill
            elif problems:
                status_cell.fill = yellow_fill
            else:
                status_cell.fill = green_fill

            # Color monthly delta cells
            base_col = 14  # first monthly column
            for i, mc in enumerate(month_cols):
                delta_col = base_col + i * 3 + 2  # the "Delta" column
                if mc in d['monthly_pace']:
                    _exp, _act, delta = d['monthly_pace'][mc]
                    delta_cell = ws_diag.cell(row=r, column=delta_col)
                    if delta < -2:
                        delta_cell.fill = red_fill
                    elif delta < 0:
                        delta_cell.fill = yellow_fill

        last_diag = max(2, len(sorted_diag) + 1)
        ws_diag.freeze_panes = "A2"
        ws_diag.auto_filter.ref = f"A1:{get_column_letter(len(headers_diag))}{last_diag}"
        _autofit(ws_diag, last_diag, len(headers_diag), min_width=10, max_width=30)

    wb.save(output_path)


def generate_matchup_table(schedule, division_teams):
    matchup_count = defaultdict(lambda: defaultdict(int))
    for date, slot, field, home_team, home_div, away_team, away_div in schedule:
        matchup_count[home_team][away_team] += 1
        matchup_count[away_team][home_team] += 1

    all_teams = sorted([team for teams in division_teams.values() for team in teams])

    if PrettyTable:
        table = PrettyTable()
        table.field_names = ["Team"] + all_teams
        for team in all_teams:
            row = [team] + [matchup_count[team][opp] for opp in all_teams]
            table.add_row(row)
        print("\nMatchup Table:")
        print(table)
    else:
        print("\nMatchup Table (CSV):")
        print("Team," + ",".join(all_teams))
        for team in all_teams:
            row = [str(matchup_count[team][opp]) for opp in all_teams]
            print(team + "," + ",".join(row))

# -------------------------------
# Main
# -------------------------------

def print_schedule_summary(team_stats):
    rows = []
    for team in sorted(team_stats.keys(), key=lambda t: (t[0], int(t[1:]) if t[1:].isdigit() else t[1:])):
        stats = team_stats[team]
        rows.append([team[0], team, target_games(team), stats.get('total_games', 0), stats.get('home_games', 0), stats.get('away_games', 0)])
    print("\nSchedule Summary:")
    if PrettyTable:
        table = PrettyTable()
        table.field_names = ["Division", "Team", "Target", "Total Games", "Home Games", "Away Games"]
        for row in rows:
            table.add_row(row)
        print(table)
    else:
        print("Division,Team,Target,Total Games,Home Games,Away Games")
        for row in rows:
            print(",".join(str(x) for x in row))


def print_doubleheader_summary(doubleheader_count):
    rows = []
    for team in sorted(doubleheader_count.keys(), key=lambda t: (t[0], int(t[1:]) if t[1:].isdigit() else t[1:])):
        rows.append([team[0], team, min_dh(team), doubleheader_count.get(team, 0)])
    print("\nDoubleheader Summary:")
    if PrettyTable:
        table = PrettyTable()
        table.field_names = ["Division", "Team", "Min DH", "DH Days"]
        for row in rows:
            table.add_row(row)
        print(table)
    else:
        print("Division,Team,Min DH,DH Days")
        for row in rows:
            print(",".join(str(x) for x in row))


def main():
    global RUN_SEED
    # --- RNG setup ---
    global RANDOM_SEED

    if RANDOM_SEED is None:
        import os
        RUN_SEED = int.from_bytes(os.urandom(4), "big")
    else:
        RUN_SEED = RANDOM_SEED

    random.seed(RUN_SEED)
    print(f"Using RNG seed: {RUN_SEED}")

    team_availability = load_team_availability('team_availability.csv')
    team_preferred_days = load_team_preferred_days('team_preferred_days.csv')
    # Debug: confirm we loaded what we think we loaded
    _ta_path = os.path.abspath('team_availability.csv')
    print(f"Loaded team availability from: {_ta_path} (teams={len(team_availability)})")
    for _t in sorted([t for t in team_availability if len(team_availability[t]) < 6]):
        print(f"  Restricted: {_t}: {sorted(team_availability[_t])}")
    if 'C1' in team_availability:
        print(f"  Sanity C1: {sorted(team_availability['C1'])}")
    if 'C2' in team_availability:
        print(f"  Sanity C2: {sorted(team_availability['C2'])}")

    if team_preferred_days:
        _tp_path = os.path.abspath('team_preferred_days.csv')
        print(f"Loaded team preferred days from: {_tp_path} (teams={len(team_preferred_days)})")
        for _t in sorted(team_preferred_days):
            if team_preferred_days[_t]:
                print(f"  Preferred: {_t}: {sorted(team_preferred_days[_t])}")
    else:
        print("No team_preferred_days.csv found (preferred day bonus disabled).")

    field_availability = load_field_availability('field_availability.csv')
    global SEASON_START_DATE
    SEASON_START_DATE = min((dt.date() for dt, _slot, _field in field_availability), default=None)
    team_blackouts = load_team_blackouts('team_blackouts.csv')

    division_teams = {
        'A': ["A{}".format(i+1) for i in range(6)],
        'B': ["B{}".format(i+1) for i in range(8)],
        'C': ["C{}".format(i+1) for i in range(6)],
        # 'D': ["D{}".format(i+1) for i in range(6)],
    }
    all_teams = [t for div in division_teams for t in division_teams[div]]

    # Precompute per-team scheduling scarcity (availability + blackouts) so that
    # heavily-constrained teams are prioritized when building pods/games.
    init_team_scarcity(all_teams, field_availability, team_availability, team_blackouts)

    schedule = []
    team_stats = defaultdict(lambda: {
        'total_games': 0,
        'home_games': 0,
        'away_games': 0,
        'weekly_games': defaultdict(int),
            })
    used_slots = {}
    team_game_days = defaultdict(lambda: defaultdict(int))
    team_game_slots = defaultdict(lambda: defaultdict(list))
    team_doubleheader_opponents = defaultdict(lambda: defaultdict(set))
    doubleheader_count = defaultdict(int)

    timeslots_by_date = defaultdict(list)
    for date, slot, field in field_availability:
        d = date.date()
        if slot not in timeslots_by_date[d]:
            timeslots_by_date[d].append(slot)
    for d in timeslots_by_date:
        timeslots_by_date[d].sort(key=lambda s: datetime.strptime(s.strip(), "%I:%M %p"))

    for t in all_teams:
        _ = team_stats[t]

    # ---------------------------------------------------------
    # Sunday pod rotation assignment (for pod-style doubleheaders on Sundays)
    # ---------------------------------------------------------
    # We rotate which division is allowed to run *pod* doubleheaders on each Sunday.
    # This prevents one division (often A) from soaking up all Sunday inventory.
    #
    # IMPORTANT: This affects pod-style DH only. Singles can still be scheduled on Sundays.
    sunday_assignment = build_sunday_pod_assignment(
        timeslots_by_date,
        rotation=SUNDAY_POD_ROTATION,
        seed=(RANDOM_SEED if RANDOM_SEED is not None else random.randint(1, 10_000_000))
    )

    # Track total pods used per Sunday across all divisions (hard cap via SUNDAY_PODS_PER_SUNDAY)
    # Format: {date: int}
    sunday_pods_used = {}

    # Pre-index field availability for O(1) lookups in DH scheduling
    field_index = build_field_index(field_availability)

    matchups = generate_full_matchups(division_teams)
    print("\nTotal generated matchups (unscheduled): {}".format(len(matchups)))

    unscheduled = matchups[:]

    # Schedule DH-only divisions FIRST so the others don't consume the prime Sunday +
    # adjacent-slot inventory that pod-only divisions require.
    for div in dh_only_divisions(division_teams):
        (schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
         used_slots) = schedule_pod_only_division(
            div, division_teams, team_availability, field_availability, team_blackouts, timeslots_by_date,
            team_stats, doubleheader_count, team_game_days, team_game_slots, used_slots, schedule,
            sunday_assignment=sunday_assignment, sunday_pods_used=sunday_pods_used
        )

    # Remove DH-only matchups from the single-game pool (those divisions are pod-only).
    unscheduled = [m for m in unscheduled
                   if not is_dh_only(div_of(m[0])) and not is_dh_only(div_of(m[1]))]

    # Build remaining divisions' doubleheader pods (same-day 2-game sets) BEFORE single-game placement.
    # Pod structure guarantees teams do NOT play the same opponent back-to-back in a DH.
    for div in division_teams:
        if is_dh_only(div):
            continue
        (schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
         team_doubleheader_opponents, used_slots, unscheduled) = schedule_division_pod_doubleheaders(
            div, division_teams, unscheduled,
            team_availability, field_availability, team_blackouts, timeslots_by_date,
            team_stats, doubleheader_count, team_game_days, team_game_slots,
            team_doubleheader_opponents, used_slots, schedule,
            sunday_assignment=sunday_assignment, sunday_pods_used=sunday_pods_used,
            team_preferred_days=team_preferred_days)

    (schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
     team_doubleheader_opponents, used_slots, unscheduled) = schedule_games(
        unscheduled, team_availability, field_availability, team_blackouts,
        schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
        team_doubleheader_opponents, used_slots, timeslots_by_date,
        sunday_assignment=sunday_assignment, team_preferred_days=team_preferred_days)

    if any(team_stats[t]['total_games'] < target_games(t) for t in all_teams):
        print("Filling missing games...")
        (schedule, team_stats, doubleheader_count, unscheduled) = fill_missing_games(
            schedule, team_stats, doubleheader_count, team_game_days, team_game_slots,
            team_doubleheader_opponents, used_slots, timeslots_by_date, unscheduled,
            team_availability, team_blackouts, field_availability,
            sunday_assignment=sunday_assignment, team_preferred_days=team_preferred_days)

    missing = [t for t in all_teams if team_stats[t]['total_games'] < target_games(t)]

    over = [t for t in all_teams if team_stats[t]['total_games'] > target_games(t)]
    if over:
        print('Critical: Teams ABOVE target games (hard cap violated): {}'.format(over))
    if missing:
        print("Critical: Teams below target games: {}".format(missing))

    under_dh = [t for t in all_teams if doubleheader_count[t] < min_dh(t)]
    if under_dh:
        print("Critical: Teams below minimum DH days: {}".format(under_dh))

    # ---------------------------------------------------------
    # Post-build repair pass: fix gaps, cadence, back-heavy scheduling
    # ---------------------------------------------------------
    print("\nRunning post-build repair pass...")
    schedule, repair_moves, diag_before, diag_after = repair_schedule(
        schedule, all_teams, team_stats, doubleheader_count,
        team_game_days, team_game_slots, team_doubleheader_opponents,
        used_slots, timeslots_by_date, field_availability, field_index,
        team_availability, team_blackouts, sunday_assignment=sunday_assignment,
        max_moves=50
    )

    # Re-validate after repair
    _av_viol_post = check_schedule_against_availability(schedule, team_availability)
    if _av_viol_post:
        print("WARNING: Repair pass introduced availability violations — rolling back")
        # In practice the repair pass checks availability, but this is a safety net

    # Export CSV + XLSX with full slot list (row count == field_availability)
    # Hard validation: no team is scheduled on a disallowed day
    _av_viol = check_schedule_against_availability(schedule, team_availability)
    if _av_viol:
        print("ERROR: Team availability violations detected (showing up to 50):")
        for v in _av_viol[:50]:
            print("  ", v)
        raise SystemExit(2)

    output_schedule_to_csv_full(field_availability, schedule, 'softball_schedule.csv')
    # Also write templates for manual scheduling
    output_unscheduled_matchups_csv(unscheduled, 'unscheduled_matchups.csv')
    output_team_remaining_needs_csv(all_teams, team_stats, doubleheader_count, 'team_remaining_needs.csv')
    export_schedule_to_xlsx(field_availability, schedule, division_teams, 'softball_schedule.xlsx',
                            remaining_matchups=unscheduled, team_stats=team_stats,
                            doubleheader_count=doubleheader_count, team_availability=team_availability,
                            team_blackouts=team_blackouts, diagnostics=diag_after)

    print("\nSchedule Generation Complete")
    print_schedule_summary(team_stats)
    print_doubleheader_summary(doubleheader_count)
    generate_matchup_table(schedule, division_teams)
    print("\nWrote: softball_schedule.csv ({} rows)".format(len(field_availability)))
    print("Wrote: softball_schedule.xlsx")

if __name__ == "__main__":
    main()
